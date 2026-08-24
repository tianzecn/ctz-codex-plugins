#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""按体检结果清洗成规范分析表，并留下可审计的痕迹。

它不只是吐出一张干净表。只给结果的清洗是黑箱 —— 三个月后没人说得清
第 47 行为什么不见了，包括做清洗的人自己。

所以它产出三样东西：
    1. 清洗后的数据（.xlsx 或 .csv）
    2. 逐步的清洗日志：每一步影响了多少行、为什么
    3. 一份等价的 pandas 脚本，别人可以读、可以改、可以复现

目标形态是 Tidy Data（Wickham）：每个变量一列，每条观测一行。

用法
----
    python3 clean_table.py <文件> [--sheet 名称] [--out 输出路径]
                                  [--dedup] [--dry-run] [--no-script]

    --dedup     去掉完全重复的行（默认保留，因为重复可能是真实业务）
    --dry-run   只报告会做什么，不写文件
    --no-script 不生成审计脚本

依赖：只需要 openpyxl（读写 .xlsx 时）。不用 pandas。
生成的审计脚本用 pandas，那是给人看和用的，与本工具的运行无关。
"""

from __future__ import annotations

import argparse
import csv
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

sys.path.insert(0, str(Path(__file__).parent))
from profile_table import (  # noqa: E402
    profile, parse_number, _norm, _is_blank, _is_null_token,
    _merge_anchor_map, read_grid, TableProfile, FULLWIDTH_SPACE,
    NULL_TOKENS, SERIAL_MIN, SERIAL_MAX, get_column_letter, HAS_OPENPYXL,
)

from datetime import datetime, date, timedelta

EXCEL_EPOCH = datetime(1899, 12, 30)      # 1900 日期系统的实际原点（含闰年 bug 补偿）


@dataclass
class Step:
    name: str
    detail: str
    rows_before: int
    rows_after: int

    @property
    def delta(self) -> int:
        return self.rows_after - self.rows_before


@dataclass
class CleanResult:
    columns: list[str] = field(default_factory=list)
    rows: list[list[Any]] = field(default_factory=list)
    steps: list[Step] = field(default_factory=list)
    col_actions: dict[str, list[str]] = field(default_factory=dict)
    unresolved: list[str] = field(default_factory=list)
    source_rows: list[int] = field(default_factory=list)   # 每行对应的原始行号


def _to_iso(raw: Any) -> tuple[Any, bool]:
    """日期归一成 ISO 字符串。返回 (值, 是否成功)。"""
    if isinstance(raw, datetime):
        return raw.date().isoformat(), True
    if isinstance(raw, date):
        return raw.isoformat(), True
    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        if SERIAL_MIN <= float(raw) <= SERIAL_MAX and float(raw).is_integer():
            return (EXCEL_EPOCH + timedelta(days=int(raw))).date().isoformat(), True
        return raw, False
    s = _norm(raw)
    if not s:
        return None, True
    # 逐个尝试常见写法
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y年%m月%d日",
                "%Y-%m", "%Y/%m", "%Y年%m月", "%Y%m%d"):
        try:
            d = datetime.strptime(s, fmt)
            return (d.date().isoformat() if "%d" in fmt
                    else d.date().isoformat()[:7]), True
        except ValueError:
            continue
    return s, False       # 认不出来的原样保留，并记进 unresolved


def clean(path: Path, sheet: str | None = None, dedup: bool = False) -> tuple[CleanResult, TableProfile]:
    p = profile(path, sheet)
    grid, meta = read_grid(path, sheet)
    res = CleanResult()

    if not p.columns or p.n_data_rows == 0:
        res.unresolved.append("没有识别出明细数据，无法清洗")
        return res, p

    amap = _merge_anchor_map(p.merged_ranges)
    width = len(p.columns)

    def cell(row: int, ci: int) -> Any:
        r = grid[row - 1] if row - 1 < len(grid) else []
        v = r[ci] if ci < len(r) else None
        if _is_blank(v) and (row, ci + 1) in amap:
            ar, ac = amap[(row, ci + 1)]
            src = grid[ar - 1] if ar - 1 < len(grid) else []
            v = src[ac - 1] if ac - 1 < len(src) else None
        return v

    # ── 步骤 1：定位表头与数据区 ─────────────────────────────────
    res.columns = [c.name for c in p.columns]
    all_span = list(range(p.data_start_row, p.data_end_row + 1))
    res.steps.append(Step(
        "定位表头",
        f"表头在第 {p.header_rows} 行"
        + (f"（{len(p.header_rows)} 级，已拍平）" if p.multi_level_header else "")
        + f"；标题/说明行 {p.title_rows or '无'} 已跳过",
        len(grid), len(all_span),
    ))

    # ── 步骤 2：排除汇总行与脚注行 ───────────────────────────────
    drop = {r["row"] for r in p.summary_rows} | {r["row"] for r in p.footnote_rows}
    kept = [r for r in all_span
            if r not in drop and not all(_is_blank(v) for v in grid[r - 1])]
    if drop:
        labels = "、".join(
            f"第{r['row']}行「{r['label'] or '?'}」" for r in p.summary_rows)
        res.steps.append(Step(
            "排除汇总行", f"{labels}；脚注 {len(p.footnote_rows)} 行",
            len(all_span), len(kept)))
    else:
        res.steps.append(Step("排除汇总行", "未发现汇总行", len(all_span), len(kept)))

    # ── 步骤 3：合并单元格 ffill ────────────────────────────────
    ffill_cols = sorted({c for (_r, c) in amap
                         if any(r == _r for r in kept)})
    if ffill_cols:
        names = "、".join(get_column_letter(c) for c in ffill_cols)
        res.steps.append(Step(
            "恢复合并单元格", f"{names} 列的空值由合并区左上角回填（等价于 ffill）",
            len(kept), len(kept)))
        for c in ffill_cols:
            res.col_actions.setdefault(
                p.columns[c - 1].name if c - 1 < width else get_column_letter(c),
                []).append("ffill 恢复分组标签")

    # ── 步骤 4：逐列按真实类型转换 ──────────────────────────────
    out_rows: list[list[Any]] = []
    disguised_hits = 0
    ws_hits = 0
    num_fail: dict[str, list[str]] = {}
    date_fail: dict[str, list[str]] = {}

    for r in kept:
        row_out: list[Any] = []
        for c in p.columns:
            raw = cell(r, c.index)
            name = c.name

            # 缺失值归一
            s = _norm(raw)
            if isinstance(raw, str) and _is_null_token(s):
                disguised_hits += 1
                row_out.append(None)
                continue
            if s == "":
                row_out.append(None)
                continue

            if isinstance(raw, str) and (raw != raw.strip() or FULLWIDTH_SPACE in raw):
                ws_hits += 1

            if c.inferred_kind == "数值":
                v, _marks = parse_number(raw)
                if v is None:
                    num_fail.setdefault(name, []).append(s)
                    row_out.append(s)          # 转不了的保留原文，不静默变 None
                else:
                    row_out.append(v)
            elif c.inferred_kind == "日期":
                v, ok = _to_iso(raw)
                if not ok:
                    date_fail.setdefault(name, []).append(s)
                row_out.append(v)
            else:
                row_out.append(s)              # 文本：去首尾+全角空格
        out_rows.append(row_out)

    for c in p.columns:
        acts = res.col_actions.setdefault(c.name, [])
        if c.inferred_kind == "数值":
            acts.append("文本→数值（去千分位/货币符号/单位后缀/括号负数）")
        elif c.inferred_kind == "日期":
            acts.append("日期→ISO 格式")
        else:
            acts.append("去首尾空格与全角空格")

    res.steps.append(Step(
        "类型转换",
        f"数值列去千分位与单位后缀；日期列统一为 ISO；"
        f"{disguised_hits} 个伪装缺失值（—、N/A、无 等）归一为真空；"
        f"{ws_hits} 处首尾/全角空格已清除",
        len(kept), len(out_rows),
    ))

    for name, vals in num_fail.items():
        uniq = sorted(set(vals))[:5]
        res.unresolved.append(
            f"「{name}」列有 {len(vals)} 个值转不成数字，已原样保留："
            f"{uniq} —— 去看看它们是什么，不要放过")
    for name, vals in date_fail.items():
        uniq = sorted(set(vals))[:5]
        res.unresolved.append(
            f"「{name}」列有 {len(vals)} 个日期认不出格式，已原样保留：{uniq}")

    # ── 步骤 5：去重（默认不做）──────────────────────────────────
    src = list(kept)
    if dedup:
        seen: set[tuple] = set()
        keep_idx: list[int] = []
        for i, row in enumerate(out_rows):
            key = tuple(_norm(v) for v in row)
            if key in seen:
                continue
            seen.add(key)
            keep_idx.append(i)
        before = len(out_rows)
        out_rows = [out_rows[i] for i in keep_idx]
        src = [src[i] for i in keep_idx]
        res.steps.append(Step("去重", "删除完全重复的行", before, len(out_rows)))
    elif p.duplicate_rows:
        rows_s = "、".join(str(d["row"]) for d in p.duplicate_rows)
        res.steps.append(Step(
            "去重", f"发现 {len(p.duplicate_rows)} 行完全重复（第 {rows_s} 行）"
                    f"但**未删除** —— 重复可能是真实业务。要删加 --dedup",
            len(out_rows), len(out_rows)))

    res.rows = out_rows
    res.source_rows = src
    return res, p


# ── 生成审计脚本 ──────────────────────────────────────────────────────
def emit_script(p: TableProfile, res: CleanResult, src: Path, dedup: bool) -> str:
    drop_rows = sorted({r["row"] for r in p.summary_rows}
                       | {r["row"] for r in p.footnote_rows})
    header_arg = ([r - 1 for r in p.header_rows]
                  if p.multi_level_header else p.header_rows[0] - 1)
    reader = "pd.read_csv" if src.suffix.lower() in (".csv", ".tsv", ".txt") else "pd.read_excel"
    num_cols = [c.name for c in p.columns if c.inferred_kind == "数值"]
    date_cols = [c.name for c in p.columns if c.inferred_kind == "日期"]
    text_cols = [c.name for c in p.columns if c.inferred_kind == "文本"]
    ffill_cols = [c.name for c in p.columns
                  if "ffill 恢复分组标签" in res.col_actions.get(c.name, [])]

    return f'''#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""{src.name} 的清洗流程（自动生成，请审阅后使用）

这份脚本是清洗过程的可读记录。每一步都能被检查、修改、复现。
生成依据来自表结构体检，但**业务判断需要你确认**，尤其是标了 ⚠ 的地方。
"""
import pandas as pd

SRC = r"{src.name}"

# ── 1. 读取：表头在第 {p.header_rows} 行（1-based）────────────────
df = {reader}(SRC, header={header_arg})

# 多级表头会读成 MultiIndex 列名，先拍平成单层，
# 否则 df["一季度 / 1月"] 取不到列（pandas 里它是 ("一季度","1月")）
if isinstance(df.columns, pd.MultiIndex):
    df.columns = [
        " / ".join(dict.fromkeys(
            str(x).strip() for x in tup
            if str(x).strip() and str(x) != "nan"
            and not str(x).startswith("Unnamed")
        )) or f"col{{i}}"
        for i, tup in enumerate(df.columns)
    ]

# ── 2. 排除非明细行 ───────────────────────────────────────────
# 原表第 {drop_rows} 行（1-based）是汇总/小计/脚注，不是明细数据。
# ⚠ 排除之前，先拿它们和你的计算结果对账 —— 那是免费的校验和。
DROP_EXCEL_ROWS = {drop_rows}
_offset = {max(p.header_rows) if p.header_rows else 1}   # 表头占掉的行数
df = df.drop(index=[r - _offset - 1 for r in DROP_EXCEL_ROWS
                    if 0 <= r - _offset - 1 < len(df)], errors="ignore")

# 整行为空的行（原表里的分隔空行）——必须在按位置删完之后再做，
# 否则 index 会错位，把不该删的行删掉
df = df.dropna(how="all").reset_index(drop=True)

# ── 3. 恢复合并单元格造成的空值 ───────────────────────────────
# ⚠ 这些列的空值不是缺失，是合并单元格的显示效果。
#    绝对不要 dropna，那会把每个分组除首行外的所有行删掉。
for col in {ffill_cols}:
    if col in df.columns:
        df[col] = df[col].ffill()

# ── 4. 缺失值归一 ─────────────────────────────────────────────
NA_TOKENS = {sorted(t for t in NULL_TOKENS if t)!r}
df = df.replace(NA_TOKENS, pd.NA)

# ── 5. 文本型数字转数值 ───────────────────────────────────────
def to_num(s):
    return pd.to_numeric(
        s.astype(str)
         .str.replace("\\u3000", "", regex=False)
         .str.strip()
         .str.replace(r"^\\((.*)\\)$", r"-\\1", regex=True)      # 会计括号负数
         .str.replace(r"[,，¥￥$€£]", "", regex=True)
         .str.replace(r"(万元|亿元|千元|万|亿|元|个|人|次|件|台|笔|单)$", "", regex=True),
        errors="coerce")

for col in {num_cols}:
    if col in df.columns:
        before_na = df[col].isna().sum()
        df[col] = to_num(df[col])
        lost = df[col].isna().sum() - before_na
        if lost:
            print(f"⚠ {{col}}：{{lost}} 个值转不成数字，已变成 NaN —— 去查是什么")

# ── 6. 日期统一 ───────────────────────────────────────────────
for col in {date_cols}:
    if col in df.columns:
        df[col] = pd.to_datetime(df[col], errors="coerce", format="mixed")
        # ⚠ 不要用 dayfirst 去猜 03/04 是三月四号还是四月三号，猜错不报错。去问。

# ── 7. 文本去空格（含全角）───────────────────────────────────
for col in {text_cols}:
    if col in df.columns:
        df[col] = (df[col].astype(str)
                          .str.replace("\\u3000", "", regex=False)
                          .str.strip()
                          .replace("nan", pd.NA))

{"# ── 8. 去重 ───────────────────────────────────────────────────" if dedup else "# ── 8. 去重（本次未启用）────────────────────────────────────"}
{"df = df.drop_duplicates()" if dedup else "# ⚠ 发现过完全重复的行但未删除。重复可能是真实业务（同一天两笔同额订单），"}
{"" if dedup else "#    也可能是粘贴事故。确认之后再决定：df = df.drop_duplicates()"}

# ── 校验：行数必须能解释 ──────────────────────────────────────
print(f"清洗后 {{len(df)}} 行 × {{len(df.columns)}} 列")
assert len(df) == {len(res.rows)}, "行数与清洗工具的结果不一致，两边对一下"

df.to_excel("cleaned.xlsx", index=False)
'''


# ── 写文件 ────────────────────────────────────────────────────────────
def write_out(res: CleanResult, out: Path) -> None:
    if out.suffix.lower() in (".csv", ".tsv"):
        delim = "\t" if out.suffix.lower() == ".tsv" else ","
        with out.open("w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f, delimiter=delim)
            w.writerow(res.columns)
            w.writerows(res.rows)
        return
    if not HAS_OPENPYXL:
        raise SystemExit("写 xlsx 需要 openpyxl，或把 --out 改成 .csv")
    from openpyxl import Workbook
    from openpyxl.styles import Font
    wb = Workbook()
    ws = wb.active
    ws.title = "cleaned"
    ws.append(res.columns)
    for c in ws[1]:
        c.font = Font(bold=True)
    for row in res.rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    wb.save(out)


def render(res: CleanResult, p: TableProfile, out: Path | None, dedup: bool) -> str:
    L: list[str] = []
    add = L.append
    bar = "─" * 68
    add(bar)
    add("清洗")
    add(bar)
    add("")
    add("【每一步做了什么】")
    for i, s in enumerate(res.steps, 1):
        d = f"{s.rows_before} → {s.rows_after}"
        if s.delta:
            d += f"（{s.delta:+d} 行）"
        add(f"  {i}. {s.name}   {d}")
        add(f"     {s.detail}")

    add("")
    add("【逐列处理】")
    for name, acts in res.col_actions.items():
        add(f"  {name}：{'；'.join(acts)}")

    if res.unresolved:
        add("")
        add("【没能自动解决 —— 需要你看一眼】")
        for u in res.unresolved:
            add(f"  ? {u}")

    add("")
    add(f"结果：{len(res.rows)} 行 × {len(res.columns)} 列")
    if out:
        add(f"已写入：{out}")
    add("")
    add(bar)
    add("清洗完不等于数字对了。接着跑 verify_numbers.py，")
    add("拿清洗后的明细去对原表的合计行 —— 那是最便宜的一次验证。")
    add(bar)
    return "\n".join(L)


def main() -> None:
    ap = argparse.ArgumentParser(description="清洗成规范分析表，并留下可审计的痕迹")
    ap.add_argument("file")
    ap.add_argument("--sheet", default=None)
    ap.add_argument("--out", default=None, help="输出路径，默认 <原名>_cleaned.xlsx")
    ap.add_argument("--dedup", action="store_true", help="删除完全重复的行")
    ap.add_argument("--dry-run", action="store_true", help="只报告，不写文件")
    ap.add_argument("--no-script", action="store_true", help="不生成审计脚本")
    a = ap.parse_args()

    src = Path(a.file).expanduser()
    if not src.exists():
        sys.exit(f"文件不存在：{src}")

    res, p = clean(src, a.sheet, a.dedup)
    if not res.rows and res.unresolved:
        print("\n".join(res.unresolved))
        sys.exit(1)

    out = None
    if not a.dry_run:
        out = Path(a.out).expanduser() if a.out else src.with_name(src.stem + "_cleaned.xlsx")
        write_out(res, out)
        if not a.no_script:
            sp = out.with_name(out.stem + "_recipe.py")
            sp.write_text(emit_script(p, res, src, a.dedup), encoding="utf-8")
            print(f"审计脚本：{sp}")

    print(render(res, p, out, a.dedup))


if __name__ == "__main__":
    main()
