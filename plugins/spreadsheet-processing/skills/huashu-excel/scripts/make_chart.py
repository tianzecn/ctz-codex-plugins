#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""图表：先判断该用什么图，再生成 Excel 原生图表。

它做两件事，第一件比第二件重要：

  --suggest   看懂数据的结构，推荐图表类型，并说明为什么、有什么风险
  生成         按推荐（或你的指定）产出图表，自动执行诚实性约束

图表类型的选择依据来自两处：
  · Zelazny《用图表说话》—— 先确定信息，再让图表形式匹配比较类型
  · Cleveland & McGill (1984) 图形感知实验 —— 视觉编码的精度阶梯：
      位置(共同基线) > 位置(非对齐) > 长度/方向/角度 > 面积 > 体积/曲率 > 明暗/色饱和
    位置判断比长度准 1.4–2.5 倍，比角度准 1.96 倍。
    所以默认给条形图而不是饼图——这是有测量结果的，不是审美偏好。

自动执行的诚实性约束（见 references/charts.md）：
  · 柱/条形图 Y 轴强制从 0（它编码长度，截断即失真）
  · 折线图不强制从 0（它编码位置和斜率，强行归零会压平真实波动）
  · 类别超过阈值时拒绝生成饼图，并说明原因
  · 检测时间序列末期是否不完整，提示标注

用法
----
    python3 make_chart.py <文件> --suggest
    python3 make_chart.py <文件> --x 门店 --y 销售额 --title "..." [--type bar]
                                 [--sheet 名称] [--out 输出.xlsx]

依赖：openpyxl。不用 pandas、不用绘图库、不联网。
"""

from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

sys.path.insert(0, str(Path(__file__).parent))
from profile_table import (  # noqa: E402
    profile, parse_number, _norm, _is_blank, _merge_anchor_map,
    read_grid, TableProfile, HAS_OPENPYXL,
)

# 超过这么多类别，图会变成噪音（见 charts.md）
MAX_CATEGORIES = 15
# 饼图的绝对上限。超过就拒绝，因为角度编码精度垫底
PIE_MAX = 3
# 时间序列点数超过这个值用折线，否则用柱形
LINE_THRESHOLD = 8
# 色盲友好的分类色板（Wong 色板系）。约 8% 的男性有红绿色觉障碍，
# 而红绿对比恰恰是最常用的选择。这套以蓝橙为主轴，避开红绿对立。
SAFE_PALETTE = ["0173B2", "DE8F05", "029E73", "CC78BC", "CA9161", "56B4E9"]


@dataclass
class Suggestion:
    chart: str
    reason: str
    x: str = ""
    y: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    title_hint: str = ""


# ── 判断列的角色 ──────────────────────────────────────────────────────
def classify_columns(p: TableProfile) -> dict[str, list]:
    """把列分成：时间列、分类列（低基数文本）、数值列、高基数文本（不可用作维度）。"""
    roles: dict[str, list] = {"time": [], "category": [], "measure": [], "id_like": []}
    n = max(p.n_data_rows, 1)
    for c in p.columns:
        live = c.n_total - c.n_missing
        if live == 0:
            continue
        if c.inferred_kind == "日期":
            roles["time"].append(c)
        elif c.inferred_kind == "数值":
            # 唯一值几乎等于行数、且是整数序列 → 更像编号不是度量
            if c.n_unique == n and c.stats and c.stats.get("iqr", 0) == 0:
                roles["id_like"].append(c)
            else:
                roles["measure"].append(c)
        else:
            # 文本：基数低才能当维度
            if c.n_unique <= MAX_CATEGORIES and c.n_unique > 1:
                roles["category"].append(c)
            elif c.n_unique >= n * 0.9:
                roles["id_like"].append(c)
            else:
                roles["category"].append(c)      # 基数偏高但仍可分组，会警告
    return roles


def suggest(p: TableProfile) -> list[Suggestion]:
    """按 Zelazny 的「比较类型 → 图表形式」推荐，附带风险提示。"""
    roles = classify_columns(p)
    out: list[Suggestion] = []
    times, cats, meas = roles["time"], roles["category"], roles["measure"]

    if not meas:
        return [Suggestion("无", "没有可用作数值的列，先确认类型转换是否成功")]

    m0 = meas[0]

    # ① 时间 + 数值 → 时间序列
    if times:
        t = times[0]
        n_pts = p.n_data_rows
        kind = "line" if n_pts > LINE_THRESHOLD else "col"
        s = Suggestion(
            kind,
            f"「{t.name}」是时间维度 → 时间序列比较。"
            + (f"{n_pts} 个时间点，用折线看趋势" if kind == "line"
               else f"只有 {n_pts} 个时间点，用柱形比折线更好读"),
            x=t.name, y=[m0.name],
            title_hint="标题写变化的结论，例：「三月起连续四个月下滑」",
        )
        s.warnings.append(
            "确认最后一期是否完整——本月只过了几天就画进去，会看起来像断崖下跌。"
            "不完整的期次要去掉或用浅色标注")
        if kind == "line":
            s.warnings.append("折线图 Y 轴不必从 0（它编码位置和斜率，不是长度）")
        else:
            s.warnings.append("柱形图 Y 轴必须从 0，否则会放大差异")
        out.append(s)

    # ② 分类 + 数值 → 项目比较（排名）
    if cats:
        c = cats[0]
        s = Suggestion(
            "bar",
            f"「{c.name}」是分类维度（{c.n_unique} 类）→ 项目比较。"
            f"用横向条形图并按值排序：条形用长度+共同基线编码，"
            f"是感知精度最高的一档",
            x=c.name, y=[m0.name],
            title_hint="标题写比较的结论，例：「深圳南山店领先第二名 20%」",
        )
        if c.n_unique > MAX_CATEGORIES:
            s.warnings.append(
                f"{c.n_unique} 个类别太多，图会变成噪音。"
                f"建议只画 Top 10 + 「其他」，或者改用表格")
        if len(c.name) > 6 or c.n_unique > 6:
            s.warnings.append("类别名较长或较多时，横向条形图比竖向柱形图好读")
        s.warnings.append("类别没有内在顺序（地区、渠道、产品）就按值排序，读者不用自己排")
        out.append(s)

        # 成分关系：明确不推荐饼图
        out.append(Suggestion(
            "bar",
            f"如果你想表达的是「占比/构成」，仍然用排序条形图 + 标注百分比，"
            f"不要用饼图：饼图靠角度和面积编码，在 Cleveland-McGill 的精度阶梯上"
            f"分别排第 3 和第 4，条形图的长度编码排第 2",
            x=c.name, y=[m0.name],
            warnings=[f"饼图只在 2–{PIE_MAX} 类且只需粗略印象时才可用；"
                      f"当前有 {c.n_unique} 类"],
            title_hint="标题写构成的结论，例：「华东贡献六成销售额」",
        ))

    # ③ 两个数值列 → 相关
    if len(meas) >= 2:
        out.append(Suggestion(
            "scatter",
            f"「{meas[0].name}」和「{meas[1].name}」都是数值 → 相关性比较，用散点图",
            x=meas[0].name, y=[meas[1].name],
            warnings=[
                "相关不是因果。除非做过随机对照实验，措辞停在「相关」，"
                "不要写成「带来」「提升」「导致」",
                "先看有没有离群点——一个异常点能把相关系数从 0.1 拉到 0.8",
            ],
            title_hint="标题写关系的结论，且不要暗示因果",
        ))

    # ④ 单数值列 → 分布
    out.append(Suggestion(
        "hist",
        f"只看「{m0.name}」本身的分布 → 频率分布比较。"
        f"业务数据几乎都是右偏的，直方图或箱线图能看出均值骗没骗你",
        x="", y=[m0.name],
        warnings=["如果分布是双峰，说明混了两群不同的对象，均值描述的是不存在的人"],
        title_hint="标题写分布的结论，例：「七成门店月销在 100–200 万之间」",
    ))

    # ⑤ 多个数值列 + 分类 → 分组对比
    if cats and len(meas) >= 2:
        out.append(Suggestion(
            "col",
            f"多个指标按「{cats[0].name}」对比 → 分组柱形图。"
            f"但如果两个指标量纲不同，不要用双 Y 轴",
            x=cats[0].name, y=[m.name for m in meas[:3]],
            warnings=[
                "双 Y 轴能凭空造出并不存在的相关性——调一下刻度，"
                "两条线就能贴合或分开。改用上下两张共享 X 轴的小图，"
                "或者把指标都转成指数（基期=100）",
            ],
        ))
    return out


# ── 生成图表 ──────────────────────────────────────────────────────────
def build(path: Path, sheet: str | None, x_name: str, y_names: list[str],
          ctype: str, title: str, out: Path) -> list[str]:
    if not HAS_OPENPYXL:
        raise SystemExit("生成图表需要 openpyxl")
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart, Reference, Series
    from openpyxl.chart.marker import Marker
    from openpyxl.styles import Font

    p = profile(path, sheet)
    grid, _ = read_grid(path, sheet)
    amap = _merge_anchor_map(p.merged_ranges)
    notes: list[str] = []

    def cell(row: int, ci: int) -> Any:
        r = grid[row - 1] if row - 1 < len(grid) else []
        v = r[ci] if ci < len(r) else None
        if _is_blank(v) and (row, ci + 1) in amap:
            ar, ac = amap[(row, ci + 1)]
            src = grid[ar - 1] if ar - 1 < len(grid) else []
            v = src[ac - 1] if ac - 1 < len(src) else None
        return v

    by_name = {c.name: c for c in p.columns}
    def find(nm: str):
        if nm in by_name:
            return by_name[nm]
        hits = [c for c in p.columns if nm and nm in c.name]
        if len(hits) == 1:
            return hits[0]
        raise SystemExit(f"找不到列「{nm}」。可用列：{[c.name for c in p.columns]}")

    xc = find(x_name) if x_name else None
    ycs = [find(n) for n in y_names]

    excluded = {r["row"] for r in p.summary_rows} | {r["row"] for r in p.footnote_rows}
    rows = [r for r in range(p.data_start_row, p.data_end_row + 1)
            if r not in excluded and not all(_is_blank(v) for v in grid[r - 1])]
    if excluded:
        notes.append(f"已排除 {len(excluded)} 个汇总/脚注行，图里只有明细数据")

    # 取数
    data: list[tuple[Any, list[float]]] = []
    for r in rows:
        xv = _norm(cell(r, xc.index)) if xc else str(r)
        yv = []
        for c in ycs:
            v, _ = parse_number(cell(r, c.index))
            yv.append(v if v is not None else 0.0)
        data.append((xv, yv))

    # 单系列且是分类维度 → 按值排序（charts.md：类别无内在顺序时按值排）
    if len(ycs) == 1 and xc and xc.inferred_kind == "文本":
        data.sort(key=lambda t: t[1][0], reverse=True)
        notes.append("类别按值降序排列——读者不用自己在脑子里排一遍")
    if len(data) > MAX_CATEGORIES and ctype in ("bar", "col", "pie"):
        keep = data[:10]
        rest = sum(v[0] for _k, v in data[10:])
        data = keep + [("其他", [rest])]
        notes.append(f"类别过多，已聚合为 Top 10 + 「其他」（原 {len(rows)} 类）")

    if ctype == "pie" and len(data) > PIE_MAX:
        raise SystemExit(
            f"拒绝生成饼图：当前 {len(data)} 个类别，超过 {PIE_MAX} 类。\n"
            f"原因：饼图靠角度和面积编码，在 Cleveland-McGill 的感知精度阶梯上"
            f"排第 3 和第 4，而条形图的长度编码排第 2。\n"
            f"改用 --type bar，并在标签上标注百分比——信息一样，读者读得更准。")

    # 写数据表
    wb = Workbook()
    ws = wb.active
    ws.title = "data"
    ws.append([xc.name if xc else "序号"] + [c.name for c in ycs])
    for c in ws[1]:
        c.font = Font(bold=True)
    for k, vs in data:
        ws.append([k] + vs)
    ws.freeze_panes = "A2"

    n = len(data)
    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=n + 1)
    data_ref = Reference(ws, min_col=2, max_col=1 + len(ycs), min_row=1, max_row=n + 1)

    if ctype in ("bar", "col"):
        ch = BarChart()
        ch.type = "bar" if ctype == "bar" else "col"
        ch.add_data(data_ref, titles_from_data=True)
        ch.set_categories(cats_ref)
        # 诚实性约束：柱/条形编码长度，基线必须为 0
        ax = ch.x_axis if ctype == "bar" else ch.y_axis
        ax.scaling.min = 0
        notes.append("数值轴已强制从 0 —— 柱形编码的是长度，截断基线会放大差异")
    elif ctype == "line":
        ch = LineChart()
        ch.add_data(data_ref, titles_from_data=True)
        ch.set_categories(cats_ref)
        notes.append("折线图未强制 Y 轴从 0 —— 它编码位置和斜率，"
                     "强行归零会把真实波动压平")
    elif ctype == "pie":
        ch = PieChart()
        ch.add_data(data_ref, titles_from_data=True)
        ch.set_categories(cats_ref)
    elif ctype == "scatter":
        ch = ScatterChart()
        ch.style = 13
        xref = Reference(ws, min_col=2, min_row=2, max_row=n + 1)
        for i in range(len(ycs)):
            yref = Reference(ws, min_col=2 + i, min_row=1, max_row=n + 1)
            s = Series(yref, xref, title_from_data=True)
            s.marker = Marker(symbol="circle", size=7)
            s.graphicalProperties.line.noFill = True      # 散点不连线
            ch.series.append(s)
        notes.append("散点图已关闭连线——连线会暗示时间顺序或因果")
    else:
        raise SystemExit(f"不支持的图表类型：{ctype}")

    ch.title = title or f"{ycs[0].name} by {xc.name if xc else ''}"
    if not title:
        notes.append("⚠ 没给标题，用了默认的主题式标题。"
                     "标题应该写结论（「深圳南山店领先第二名 20%」）"
                     "而不是主题（「各门店销售额」）——读者只看几秒")

    # 做工：去掉不承载信息的墨水
    ch.height, ch.width = 9, 18
    if hasattr(ch, "y_axis"):
        ch.y_axis.majorGridlines = None
    if len(ycs) == 1:
        ch.legend = None
        notes.append("单系列已去掉图例——图例会让眼睛在颜色和标签之间来回跑")

    # 色盲友好配色：避开红绿对立
    try:
        for i, s in enumerate(ch.series):
            color = SAFE_PALETTE[i % len(SAFE_PALETTE)]
            if ctype == "line":
                s.graphicalProperties.line.solidFill = color
                s.graphicalProperties.line.width = 22000      # EMU，约 1.75pt
            else:
                s.graphicalProperties.solidFill = color
                s.graphicalProperties.line.solidFill = color
        if len(ch.series) > 1:
            notes.append(f"已套用色盲友好色板（避开红绿对立，"
                         f"约 8% 的男性有红绿色觉障碍）")
    except (AttributeError, TypeError):
        pass          # 某些图表类型不支持逐系列上色，不因为配色失败而中断

    ws2 = wb.create_sheet("chart")
    ws2.add_chart(ch, "B2")
    wb.save(out)
    return notes


def render_suggestions(sugs: list[Suggestion], p: TableProfile) -> str:
    L = ["─" * 68, "图表建议", "─" * 68, ""]
    L.append("第一步不是选图表，是把要传达的信息写成一个完整的句子。")
    L.append("写不出那句话，画任何图都是在装饰数据。")
    L.append("")
    for i, s in enumerate(sugs, 1):
        L.append(f"{i}. {s.chart}"
                 + (f"   x={s.x}" if s.x else "")
                 + (f"   y={', '.join(s.y)}" if s.y else ""))
        L.append(f"   {s.reason}")
        if s.title_hint:
            L.append(f"   标题：{s.title_hint}")
        for w in s.warnings:
            L.append(f"   ⚠ {w}")
        L.append("")
    L.append("─" * 68)
    L.append("生成：make_chart.py <文件> --x <列> --y <列> --type <bar|col|line|scatter> --title \"结论\"")
    L.append("─" * 68)
    return "\n".join(L)


def main() -> None:
    ap = argparse.ArgumentParser(description="判断该用什么图，然后生成 Excel 原生图表")
    ap.add_argument("file")
    ap.add_argument("--sheet", default=None)
    ap.add_argument("--suggest", action="store_true", help="只推荐图表类型，不生成")
    ap.add_argument("--x", default="", help="维度列名")
    ap.add_argument("--y", action="append", default=[], help="数值列名，可重复")
    ap.add_argument("--type", default="bar",
                    choices=["bar", "col", "line", "pie", "scatter"])
    ap.add_argument("--title", default="", help="图表标题，写结论不写主题")
    ap.add_argument("--out", default=None)
    a = ap.parse_args()

    src = Path(a.file).expanduser()
    if not src.exists():
        sys.exit(f"文件不存在：{src}")

    if a.suggest or not a.y:
        p = profile(src, a.sheet)
        print(render_suggestions(suggest(p), p))
        if not a.suggest:
            print("\n（没有指定 --y，只做了推荐。加上 --x/--y 才会生成图表。）")
        return

    out = Path(a.out).expanduser() if a.out else src.with_name(src.stem + "_chart.xlsx")
    notes = build(src, a.sheet, a.x, a.y, a.type, a.title, out)
    print(f"已生成：{out}")
    for n in notes:
        print(f"  · {n}")


if __name__ == "__main__":
    main()
