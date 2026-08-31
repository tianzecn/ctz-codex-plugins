#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""表结构体检 —— 在动手算任何数字之前，先搞清楚这张表到底长什么样。

为什么需要它
------------
主流做法是 `pd.read_excel(f)` 然后 `df.describe()`。问题在于：pandas 读进来的
那一刻，合并单元格、单元格格式、原始类型就已经全部丢失了。之后所有的判断都
建立在已经损毁的信息上。

实测：一份典型的中国公司销售表（标题占首行、两级表头、地区列合并、金额带
千分位、尾部三行汇总），用「发现表头位置 + 清掉千分位」这种看起来很聪明的
做法去算 1 月总额，得到的结果比真值高 161%，而且零报错、零 NaN、零警告。

所以这个脚本先用 openpyxl 读原始单元格，把结构问题全部标出来，再决定怎么读。

用法
----
    python3 profile_table.py <文件> [--sheet 名称] [--json] [--max-scan 200]

支持 .xlsx / .xlsm / .csv / .tsv。CSV 没有合并单元格和格式，会自动跳过相关检查。

依赖
----
只需要 openpyxl（读 .xlsx 时）。CSV 路径连 openpyxl 都不需要，纯标准库。
不用 pandas、不用 numpy、不用 LibreOffice、不联网、不依赖任何 agent 平台特性。

这不是为了炫技：体检要在「还没决定用什么工具处理这张表」的时刻就能跑，
所以它自己必须几乎没有前置条件。
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import re
import sys
from collections import Counter
from dataclasses import dataclass, field, asdict
from datetime import date, datetime, time
from html.parser import HTMLParser
from pathlib import Path
from typing import Any

# ── 可选依赖 ──────────────────────────────────────────────────────────
try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:  # pragma: no cover
    HAS_OPENPYXL = False

    def get_column_letter(idx: int) -> str:
        s = ""
        while idx > 0:
            idx, r = divmod(idx - 1, 26)
            s = chr(65 + r) + s
        return s


# ── 中文办公表格的常见词表 ────────────────────────────────────────────
# 这些词出现在行首/首列时，强烈暗示该行不是明细数据而是汇总、小计或脚注。
SUMMARY_WORDS = (
    "合计", "总计", "小计", "总和", "累计", "汇总", "求和", "平均", "均值",
    "占比", "比例", "同比", "环比", "增长率", "总结", "总额", "共计",
    "total", "subtotal", "sum", "avg", "average", "grand total",
)
FOOTNOTE_PREFIX = ("注：", "注:", "备注", "说明：", "说明:", "数据来源", "来源：",
                   "来源:", "制表", "填表", "*", "※", "note:", "source:")

# 伪装成数据的缺失值。真实表格里这些都表示「没有」，但 pandas 会当成字符串。
NULL_TOKENS = {
    "", "-", "--", "—", "——", "/", "\\", "n/a", "na", "n.a.", "null", "none",
    "nan", "#n/a", "无", "空", "暂无", "未知", "不适用", "待定", "?", "？",
}

# 数字前后可能挂着的东西
CURRENCY = "¥￥$€£"
UNIT_SUFFIX = ("万元", "亿元", "千元", "百万", "万", "亿", "元", "个", "人", "次",
               "件", "台", "笔", "单", "天", "小时", "分钟", "%", "‰")

# Excel 1900 日期系统的合理业务区间：1990-01-01 ≈ 32874，2100-01-01 ≈ 73051
SERIAL_MIN, SERIAL_MAX = 32874, 73051

# 无分隔符的纯数字日期（Excel 序列号、YYYYMMDD）要占到这个比例，
# 才认为整列真是日期。低于它说明那些是碰巧长得像日期的编号或金额——
# 五位整数是订单号/工号/门店编号最常见的形态，八位整数则常常是
# 「19,980,605 元」这种金额。带分隔符的 2026-08-24 不受此限，它不会是钱。
AMBIGUOUS_DATE_LABELS = ("Excel日期序列号", "YYYYMMDD")
SERIAL_DATE_MIN_SHARE = 0.30

FULLWIDTH_SPACE = "　"

# ── 标识列 ────────────────────────────────────────────────────────────
# 长得像数字但不是数量的列：订单号、工号、机构编码、邮编、电话。
# 对它们做类型转换会丢前导零、把 join 键变成浮点；算均值、方差、离群点
# 则毫无意义（「预算科目编号的均值」不是一个数）。
# 注意这里没有 _code / _number——它们必须卡后缀（见 IDENT_SUFFIX_WORD），
# 否则 object_code_name（名称列）和 budget_number_of_contracts（合同数量）
# 会被当成编号列，而它们一个是文本一个是真度量。
IDENTIFIER_NAME_HINTS = (
    "zip", "postal",
    "phone", "tel_", "mobile", "sku", "isbn", "uuid", "guid",
    "编号", "编码", "代码", "号码", "工号", "单号", "卡号", "邮编", "邮政编码",
    "电话", "手机", "证号", "科目号", "机构号",
)
# 短词要卡词边界，否则 paid / bruno / memo 这类会被误判成编号列。
IDENT_SUFFIX_WORD = re.compile(r"(?:^|[\s_\-])(id|no|num|code|number)$", re.IGNORECASE)
IDENT_SUFFIX_CAMEL = re.compile(r"[a-z](Id|No|Num|Code|Number)$")


# ── 数据结构 ──────────────────────────────────────────────────────────
@dataclass
class ColumnProfile:
    index: int
    letter: str
    name: str
    declared_kind: str = ""          # 从单元格原始类型看是什么
    inferred_kind: str = ""          # 清洗之后真正是什么
    n_total: int = 0
    n_missing: int = 0
    n_unique: int = 0
    is_identifier: bool = False       # 编号/代码/邮编这类：不转换、不做统计
    identifier_reason: str = ""
    missing_disguised: dict[str, int] = field(default_factory=dict)
    issues: list[str] = field(default_factory=list)
    samples: list[str] = field(default_factory=list)
    stats: dict[str, Any] = field(default_factory=dict)


@dataclass
class TableProfile:
    path: str = ""
    sheet: str = ""
    dimensions: str = ""
    header_rows: list[int] = field(default_factory=list)
    multi_level_header: bool = False
    data_start_row: int = 0
    data_end_row: int = 0
    n_data_rows: int = 0
    title_rows: list[int] = field(default_factory=list)
    summary_rows: list[dict] = field(default_factory=list)
    footnote_rows: list[dict] = field(default_factory=list)
    merged_ranges: list[str] = field(default_factory=list)
    merged_in_data: list[dict] = field(default_factory=list)
    duplicate_rows: list[dict] = field(default_factory=list)
    hidden_rows: list[int] = field(default_factory=list)
    hidden_cols: list[str] = field(default_factory=list)
    columns: list[ColumnProfile] = field(default_factory=list)
    read_hint: dict[str, Any] = field(default_factory=dict)
    warnings: list[str] = field(default_factory=list)


# ── 小工具 ────────────────────────────────────────────────────────────
def _norm(v: Any) -> str:
    """单元格值转成用于判断的规范字符串。"""
    if v is None:
        return ""
    if isinstance(v, str):
        return v.replace(FULLWIDTH_SPACE, " ").strip()
    return str(v).strip()


def _is_blank(v: Any) -> bool:
    return _norm(v) == ""


def _is_null_token(s: str) -> bool:
    return s.strip().lower() in NULL_TOKENS


def _longest_run(rows: list[int]) -> int:
    """最长的一段连续行号有多长。用来区分「零星小计」和「连成片」。"""
    if not rows:
        return 0
    best = run = 1
    for a, b in zip(rows, rows[1:]):
        run = run + 1 if b == a + 1 else 1
        best = max(best, run)
    return best


def _fmt_rows(rows: list[int], cap: int = 8) -> str:
    """行号列表印成人能读的样子。几百个行号铺满屏幕，等于没有报告。"""
    rows = list(rows)
    if len(rows) <= cap:
        return ", ".join(map(str, rows))
    head = ", ".join(map(str, rows[:cap]))
    return f"{head} … 等 {len(rows)} 行（第 {rows[0]}–{rows[-1]} 行之间）"


def _looks_like_summary(s: str) -> bool:
    """标签看起来像汇总/小计吗。

    用「包含」而不是「开头」匹配，因为中文表格里「华东小计」「全国合计」
    这种关键词在词尾的写法是主流。代价是「合计商店」这类店名会被误判。

    这个误判方向是刻意选的：漏掉一个小计行会让求和翻倍且悄无声息；
    误判一行只是少算一行，而且报告里会把每个判定出的汇总行列出来供核对。
    错误代价不对称，就往代价小的那边错。
    """
    t = s.strip().lower()
    if not t or len(t) > 12:      # 更长的多半是句子而非标签
        return False
    return any(w in t for w in SUMMARY_WORDS)


def _looks_like_footnote(s: str) -> bool:
    t = s.strip().lower()
    return bool(t) and any(t.startswith(p) for p in FOOTNOTE_PREFIX)


NUM_RE = re.compile(r"^[+\-]?[\d,，]*\.?\d+(?:[eE][+\-]?\d+)?$")


def parse_number(raw: Any) -> tuple[float | None, list[str]]:
    """尽最大努力把一个单元格解析成数字，并报告它被什么东西污染了。

    返回 (值, 污染标签列表)。解析不出来返回 (None, 标签)。
    """
    marks: list[str] = []
    if raw is None:
        return None, marks
    if isinstance(raw, bool):
        return None, ["布尔值"]
    if isinstance(raw, (int, float)):
        if isinstance(raw, float) and math.isnan(raw):
            return None, marks
        return float(raw), marks

    s = str(raw)
    if s != s.strip() or FULLWIDTH_SPACE in s:
        marks.append("首尾空白/全角空格")
    s = s.replace(FULLWIDTH_SPACE, "").strip()
    if not s or _is_null_token(s):
        return None, marks

    # 括号负数：(123) → -123（会计惯例）
    neg = False
    if s.startswith("(") and s.endswith(")"):
        neg, s = True, s[1:-1].strip()
        marks.append("括号负数")

    for c in CURRENCY:
        if c in s:
            s = s.replace(c, "")
            marks.append("货币符号")

    pct = False
    for suf in sorted(UNIT_SUFFIX, key=len, reverse=True):
        if s.endswith(suf):
            s = s[: -len(suf)].strip()
            if suf in ("%", "‰"):
                pct = True
                marks.append("百分号")
            else:
                marks.append(f"单位后缀「{suf}」")
            break

    if "," in s or "，" in s:
        marks.append("千分位")
        s = s.replace(",", "").replace("，", "")

    s = s.lstrip("+")
    if s.startswith("+"):
        s = s[1:]

    if not NUM_RE.match(s.replace(",", "")):
        return None, marks
    try:
        val = float(s)
    except ValueError:
        return None, marks
    if neg:
        val = -val
    if pct:
        val /= 100.0
    return val, marks


# 每条：正则、标签、各捕获组的合法区间（None 表示不校验）
# 光靠形状匹配会出事：字符串 "1883.69" 完美符合「YYYY.M」的形状，
# 但 69 不是月份，它其实是个小数。所以每个模式都必须校验数值区间。
DATE_PATTERNS = [
    (re.compile(r"^(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})$"), "YYYY-M-D",
     [(1900, 2200), (1, 12), (1, 31)]),
    (re.compile(r"^(\d{4})年(\d{1,2})月(\d{1,2})日$"), "YYYY年M月D日",
     [(1900, 2200), (1, 12), (1, 31)]),
    (re.compile(r"^(\d{4})年(\d{1,2})月$"), "YYYY年M月",
     [(1900, 2200), (1, 12)]),
    (re.compile(r"^(\d{1,2})[-/](\d{1,2})[-/](\d{4})$"), "D-M-YYYY",
     [(1, 31), (1, 12), (1900, 2200)]),
    (re.compile(r"^(\d{1,2})月(\d{1,2})日$"), "M月D日（缺年份）",
     [(1, 12), (1, 31)]),
    # 只认 - 和 /：句点分隔的「YYYY.M」和小数无法区分，宁可漏认不可错认
    (re.compile(r"^(\d{4})[-/](\d{1,2})$"), "YYYY-M",
     [(1900, 2200), (1, 12)]),
    (re.compile(r"^(\d{4})(\d{2})(\d{2})$"), "YYYYMMDD",
     [(1900, 2200), (1, 12), (1, 31)]),
    (re.compile(r"^(\d{4})Q([1-4])$", re.I), "YYYYQn",
     [(1900, 2200), (1, 4)]),
]


def sniff_identifier(name: str, raws: list[Any]) -> str:
    """这一列是不是标识列（编号/代码/邮编/电话）？是就返回理由，否则空串。

    三条判据都是硬证据，宁可漏判也不误伤金额列——把一列钱错判成标识列，
    代价是它不再被检查离群值；而把一列编号错判成数量，代价是主键被静默摧毁。

    判据一：列名命中标识词表。
    判据二：存在带前导零的值。数量不会写成 `002`，编号才会。
    判据三：同列并存纯数字与字母数字混合值（`615` 与 `10E`、`536365` 与 `C489449`）。
            这种列一转数值就一半整数一半字符串，反而制造出「混合类型」。
    """
    raw_name = (name or "").strip()
    low = raw_name.lower()
    for hint in IDENTIFIER_NAME_HINTS:
        if hint in low:
            return f"列名含「{hint}」"
    # 「id」「no」这两个太短，子串匹配会误伤 paid / no.（避免把金额列当编号），
    # 所以要求它们出现在词边界上：Customer ID / order_no / customerId 都算。
    if IDENT_SUFFIX_WORD.search(raw_name) or IDENT_SUFFIX_CAMEL.search(raw_name):
        return f"列名以编号词结尾（{raw_name}）"

    leading_zero = 0
    pure_digit = 0
    alnum_mixed = 0
    live = 0
    for raw in raws:
        if isinstance(raw, str):
            s = raw.strip()
            if not s:
                continue
            live += 1
            if s.isdigit():
                pure_digit += 1
                if len(s) > 1 and s[0] == "0":
                    leading_zero += 1
            elif (s.isalnum() and any(ch.isdigit() for ch in s)
                  and any(ch.isalpha() for ch in s)):
                alnum_mixed += 1
        elif isinstance(raw, (int, float)) and not isinstance(raw, bool):
            live += 1
            if float(raw).is_integer():
                pure_digit += 1

    if leading_zero:
        return f"{leading_zero} 个值带前导零（002 这种，转数值就丢了）"
    # 两边都要占到一定比例才算——否则一列商品描述里混进几个纯数字也会中招。
    floor = max(1, int(live * 0.01))
    if pure_digit >= floor and alnum_mixed >= floor:
        return (f"纯数字({pure_digit})与字母数字({alnum_mixed})并存，"
                f"转数值只会转走一半，反而制造混合类型列")
    return ""


def sniff_date(raw: Any, allow_serial: bool = True) -> str | None:
    """认出日期长什么样。返回格式标签，认不出返回 None。

    只匹配形状是不够的——必须校验月份、日期落在合法区间，
    否则小数、编号、金额都会被误认成日期，进而让整列被判成「混合类型」。

    数值分支尤其危险：任何落在 32874–73051 的整数形状上都像序列号，
    而五位整数正是订单号、工号、门店编号最常见的形态。所以这里只做形状判断，
    「这一列到底是不是日期」由调用方按两道列级守卫决定——
    `allow_serial=False` 关掉标识列，`SERIAL_DATE_MIN_SHARE` 挡掉零星命中。
    """
    if isinstance(raw, (datetime, date, time)):
        return "真日期类型"
    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        if not allow_serial:
            return None
        if SERIAL_MIN <= float(raw) <= SERIAL_MAX and float(raw).is_integer():
            return "Excel日期序列号"
        return None
    s = _norm(raw)
    if not s:
        return None
    for pat, label, ranges in DATE_PATTERNS:
        m = pat.match(s)
        if not m:
            continue
        try:
            vals = [int(g) for g in m.groups()]
        except (TypeError, ValueError):
            continue
        if len(vals) != len(ranges):
            continue
        if all(lo <= v <= hi for v, (lo, hi) in zip(vals, ranges)):
            return label
    return None


def five_number(values: list[float]) -> dict[str, Any]:
    """Tukey 五数概括 + 抗差离群检测。

    刻意不把 mean/std 放在前面：业务数据（销售额、客单价、停留时长）几乎总是
    右偏的，均值会被少数大客户拉走。中位数和 IQR 对偏态稳健得多。
    —— John Tukey,《Exploratory Data Analysis》(1977)
    """
    if not values:
        return {}
    xs = sorted(values)
    n = len(xs)

    def q(p: float) -> float:
        if n == 1:
            return xs[0]
        pos = p * (n - 1)
        lo = int(math.floor(pos))
        hi = min(lo + 1, n - 1)
        frac = pos - lo
        return xs[lo] * (1 - frac) + xs[hi] * frac

    q1, med, q3 = q(0.25), q(0.5), q(0.75)
    iqr = q3 - q1
    lo_f, hi_f = q1 - 1.5 * iqr, q3 + 1.5 * iqr
    outliers = [x for x in xs if x < lo_f or x > hi_f]
    mean = sum(xs) / n
    var = sum((x - mean) ** 2 for x in xs) / n if n > 1 else 0.0
    std = math.sqrt(var)

    skew_note = ""
    if iqr > 0:
        # 中位数在箱体里的相对位置，偏离 0.5 越远越偏
        pos = (med - q1) / iqr
        if pos < 0.35:
            skew_note = "右偏（少数大值拉高均值，用中位数更代表典型情况）"
        elif pos > 0.65:
            skew_note = "左偏"
    if std > 0 and abs(mean - med) / std > 0.3 and not skew_note:
        skew_note = "均值与中位数明显分离，分布不对称"

    return {
        "min": xs[0], "q1": q1, "median": med, "q3": q3, "max": xs[-1],
        "iqr": iqr, "n": n,
        "outlier_count": len(outliers),
        "outlier_examples": outliers[:5],
        "mean": mean, "std": std,
        "distribution_note": skew_note,
    }


# ── 结构侦测 ──────────────────────────────────────────────────────────
def _row_signature(cells: list[Any]) -> tuple[int, int, int]:
    """返回 (非空数, 文本数, 数字数)，用于判断这行像表头还是像数据。"""
    nonempty = text = num = 0
    for c in cells:
        if _is_blank(c):
            continue
        nonempty += 1
        v, _ = parse_number(c)
        if v is not None and not isinstance(c, str):
            num += 1
        elif v is not None:
            num += 1
        else:
            text += 1
    return nonempty, text, num


def detect_header(grid: list[list[Any]], max_scan: int = 30) -> tuple[list[int], list[int], int]:
    """找出标题行、表头行、数据起始行。

    判据（按可靠性排序）：
      1. 表头行是「全文本且非空格子多」的行
      2. 它下面紧跟着的行开始出现数字
      3. 标题行的特征是「只有一两个格子有值」（通常还合并了）
    """
    _HEADER_WARNINGS.clear()
    n = min(len(grid), max_scan)
    title_rows: list[int] = []
    header_rows: list[int] = []

    sigs = [_row_signature(r) for r in grid[:n]]
    width = max((s[0] for s in sigs), default=0)
    if width == 0:
        return [], [], 1

    # 第一遍：把开头那些「只有极少格子有值」的行归为标题/说明
    i = 0
    while i < n:
        nonempty, text, num = sigs[i]
        if nonempty == 0:
            title_rows.append(i + 1)
            i += 1
            continue
        # 只有 1-2 个非空且宽度明显不足 → 标题或制表说明
        if nonempty <= max(2, width // 4) and num == 0:
            title_rows.append(i + 1)
            i += 1
            continue
        break

    # 第二遍：从这里开始找连续的「全文本」行，它们就是表头（可能多级）
    j = i
    while j < n:
        nonempty, text, num = sigs[j]
        if nonempty == 0:
            break
        # 表头特征：非空格子里几乎没有数字
        if nonempty >= max(2, width // 3) and num <= max(0, nonempty // 5):
            header_rows.append(j + 1)
            j += 1
            continue
        break

    # 这一段吃多了会静默删数据，所以宁可判少不判多：
    # 一列数值格、其余全是文本的数据行（境外票据没有税额、问卷只填了一栏）
    # 完全符合上面的表头特征，会被一路吃到下一行出现多个数字为止。
    # 两个反证据，命中任一就退回只认第一行——多认一行数据的代价是体检里多几个脏值，
    # 看得见；少认一行数据的代价是它从此不存在。
    if len(header_rows) > 1:
        why = _header_overrun_reason(grid, header_rows, width)
        if why:
            header_rows = header_rows[:1]
            j = i + 1
            _HEADER_WARNINGS.append(why)

    if not header_rows:
        # 兜底：把第一行当表头
        header_rows = [i + 1] if i < n else [1]
        j = i + 1

    return title_rows, header_rows, j + 1


# detect_header 退回单行表头时，把原因记在这里，由体检报告打出来。
_HEADER_WARNINGS: list[str] = []

MAX_HEADER_ROWS = 4  # 见过最深的多级表头是 4 层；再深基本是把数据行当成了表头


def _header_overrun_reason(
    grid: list[list[Any]], header_rows: list[int], width: int
) -> str:
    """候选表头里有没有「这根本不是表头」的证据。有就返回一句人话，没有返回空串。"""
    if len(header_rows) > MAX_HEADER_ROWS:
        return (
            f"识别到 {len(header_rows)} 行连续表头，超过 {MAX_HEADER_ROWS} 层——"
            f"多级表头极少这么深，判断是把数据行当成了表头，已退回只认第 {header_rows[0]} 行。"
            f"若这张表确实是多级表头，用 --header 手工指定。"
        )
    # 表头的每一层里，同一列不会反复出现同一个非空值；数据行会（「进项」重复 31 次）。
    for col in range(width):
        seen: dict[str, int] = {}
        for r in header_rows:
            row = grid[r - 1]
            if col >= len(row) or _is_blank(row[col]):
                continue
            key = str(row[col]).strip()
            seen[key] = seen.get(key, 0) + 1
            if seen[key] >= 3:
                return (
                    f"候选表头第 {col + 1} 列里「{key}」重复出现 {seen[key]} 次——"
                    f"表头的每一层不会自我重复，判断第 {header_rows[1]} 行起已经是数据，"
                    f"已退回只认第 {header_rows[0]} 行。"
                )
    return ""


RANGE_RE = re.compile(r"^([A-Z]+)(\d+):([A-Z]+)(\d+)$")


def _col_to_idx(letters: str) -> int:
    """A→1, B→2, ..., AA→27"""
    n = 0
    for ch in letters:
        n = n * 26 + (ord(ch) - 64)
    return n


def _merge_anchor_map(merged: list[str]) -> dict[tuple[int, int], tuple[int, int]]:
    """(行,列) → 该合并区左上角的 (行,列)。都是 1-based。"""
    amap: dict[tuple[int, int], tuple[int, int]] = {}
    for rng in merged:
        m = RANGE_RE.match(rng.strip().upper())
        if not m:
            continue
        c1, r1, c2, r2 = _col_to_idx(m.group(1)), int(m.group(2)), _col_to_idx(m.group(3)), int(m.group(4))
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                if (r, c) != (r1, c1):
                    amap[(r, c)] = (r1, c1)
    return amap


def _build_names(grid: list[list[Any]], header_rows: list[int], width: int,
                 merged: list[str] | None = None) -> list[str]:
    """多级表头拍平成列名。

    跨列合并的表头（如「一季度」横跨 E:G）只有左上角单元格有值，其余是空。
    继承必须严格限制在合并范围内 —— 无脑向右填充会把「3月」一路传给
    右边所有列，让「上半年合计」变成「上半年合计 / 3月」。
    有 openpyxl 的合并信息就用它；没有（比如 CSV）才退回向右填充。
    """
    if not header_rows:
        return [get_column_letter(c + 1) for c in range(width)]

    amap = _merge_anchor_map(merged or [])
    levels: list[list[str]] = []
    for r in header_rows:
        row = grid[r - 1] if r - 1 < len(grid) else []
        vals: list[str] = []
        for c in range(width):
            v = _norm(row[c]) if c < len(row) else ""
            if not v and (r, c + 1) in amap:
                ar, ac = amap[(r, c + 1)]
                src = grid[ar - 1] if ar - 1 < len(grid) else []
                v = _norm(src[ac - 1]) if ac - 1 < len(src) else ""
            vals.append(v)
        if not amap:
            # 没有合并信息时的兜底：向右继承
            last = ""
            for c in range(width):
                if vals[c]:
                    last = vals[c]
                elif last:
                    vals[c] = last
        levels.append(vals)

    names = []
    for c in range(width):
        parts, seen = [], set()
        for lv in levels:
            v = lv[c]
            if v and v not in seen:
                parts.append(v)
                seen.add(v)
        names.append(" / ".join(parts) if parts else get_column_letter(c + 1))
    return names


# ── 主流程 ────────────────────────────────────────────────────────────
def read_grid(path: Path, sheet: str | None) -> tuple[list[list[Any]], dict]:
    """读成二维原始值网格，同时把只有 openpyxl 才看得到的信息带出来。"""
    meta: dict[str, Any] = {"merged": [], "hidden_rows": [], "hidden_cols": [],
                            "sheet": "", "dimensions": "", "source": ""}

    if path.suffix.lower() in (".csv", ".tsv", ".txt"):
        delim = "\t" if path.suffix.lower() == ".tsv" else ","
        raw = path.read_bytes()
        for enc in ("utf-8-sig", "utf-8", "gb18030", "gbk", "latin-1"):
            try:
                text = raw.decode(enc)
                meta["encoding"] = enc
                break
            except UnicodeDecodeError:
                continue
        else:  # pragma: no cover
            text = raw.decode("utf-8", errors="replace")
            meta["encoding"] = "utf-8(replace)"
        rows = list(csv.reader(text.splitlines(), delimiter=delim))
        meta["sheet"] = path.name
        meta["source"] = "csv"
        meta["dimensions"] = f"{len(rows)} 行"
        return rows, meta

    # 扩展名是最不可信的一条线索。中国各家后台导出的 .xls 里，
    # 真正是 Excel 二进制格式的只占一部分，另一大类根本就是 HTML 表格改了个后缀。
    kind = _sniff(path)
    if kind == "ole2":
        return _read_ole2(path, sheet, meta)
    if kind == "html":
        return _read_html_table(path, meta)

    if not HAS_OPENPYXL:
        raise SystemExit("读 xlsx 需要 openpyxl：pip install openpyxl")

    wb = load_workbook(path, data_only=True, read_only=False)
    ws = wb[sheet] if sheet else wb.active
    meta["sheet"] = ws.title
    meta["all_sheets"] = wb.sheetnames
    meta["dimensions"] = ws.dimensions or ""
    meta["source"] = "xlsx"
    meta["merged"] = [str(r) for r in ws.merged_cells.ranges]
    meta["hidden_rows"] = [r for r, d in ws.row_dimensions.items() if d.hidden]
    meta["hidden_cols"] = [c for c, d in ws.column_dimensions.items() if d.hidden]

    grid = [list(row) for row in ws.iter_rows(values_only=True)]
    # 去掉尾部整行为空
    while grid and all(_is_blank(v) for v in grid[-1]):
        grid.pop()
    return grid, meta


def _sniff(path: Path) -> str:
    """看前几个字节，别信扩展名。

    返回 ole2（Excel 97-2003 二进制）/ html（表格改了个后缀）/ zip（真 xlsx）/ unknown。
    """
    with path.open("rb") as fh:
        head = fh.read(1024)
    if head[:4] == b"\xd0\xcf\x11\xe0":
        return "ole2"
    if head[:4] == b"PK\x03\x04":
        return "zip"
    probe = head[:512].lower()
    if b"<table" in probe or b"<html" in probe or probe.lstrip()[:1] == b"<":
        return "html"
    return "unknown"


def _read_ole2(path: Path, sheet: str | None, meta: dict) -> tuple[list[list[Any]], dict]:
    """Excel 97-2003 的二进制格式。openpyxl 不支持，只能靠 xlrd。"""
    try:
        import xlrd  # type: ignore
    except ImportError:
        raise SystemExit(
            f"{path.name} 是 Excel 97-2003 二进制格式（.xls），openpyxl 读不了。\n"
            f"装 xlrd 即可：pip install xlrd\n"
            f"或者用 Excel / LibreOffice 另存为 .xlsx 再跑。"
        )
    book = xlrd.open_workbook(str(path))
    ws = book.sheet_by_name(sheet) if sheet else book.sheet_by_index(0)
    meta["sheet"] = ws.name
    meta["all_sheets"] = book.sheet_names()
    meta["dimensions"] = f"{ws.nrows} 行 x {ws.ncols} 列"
    meta["source"] = "xls(ole2)"
    # xlrd 认得合并区，格式是 (行首, 行尾, 列首, 列尾) 且右开；转成 A1:B2 的写法
    meta["merged"] = [
        f"{get_column_letter(c0 + 1)}{r0 + 1}:{get_column_letter(c1)}{r1}"
        for r0, r1, c0, c1 in getattr(ws, "merged_cells", [])
    ]
    grid = [[ws.cell_value(r, c) for c in range(ws.ncols)] for r in range(ws.nrows)]
    while grid and all(_is_blank(v) for v in grid[-1]):
        grid.pop()
    return grid, meta


class _TableGrab(HTMLParser):
    """把第一张 <table> 抠成二维数组。只认 tr/td/th，忽略其余标签。"""

    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.rows: list[list[str]] = []
        self._row: list[str] | None = None
        self._cell: list[str] | None = None

    def handle_starttag(self, tag: str, attrs: list) -> None:
        if tag == "tr":
            self._row = []
        elif tag in ("td", "th") and self._row is not None:
            self._cell = []

    def handle_endtag(self, tag: str) -> None:
        if tag in ("td", "th") and self._cell is not None and self._row is not None:
            self._row.append("".join(self._cell).strip())
            self._cell = None
        elif tag == "tr" and self._row is not None:
            self.rows.append(self._row)
            self._row = None

    def handle_data(self, data: str) -> None:
        if self._cell is not None:
            self._cell.append(data)


def _read_html_table(path: Path, meta: dict) -> tuple[list[list[Any]], dict]:
    """后台导出最常见的一种伪装：HTML 表格，存成 .xls 骗 Excel 打开。"""
    raw = path.read_bytes()
    for enc in ("utf-8-sig", "utf-8", "gb18030", "utf-16"):
        try:
            text = raw.decode(enc)
            meta["encoding"] = enc
            break
        except UnicodeDecodeError:
            continue
    else:  # pragma: no cover
        text = raw.decode("utf-8", errors="replace")
        meta["encoding"] = "utf-8(replace)"
    g = _TableGrab()
    g.feed(text)
    rows = [r for r in g.rows if r]
    if not rows:
        raise SystemExit(
            f"{path.name} 的内容是 HTML 而不是 Excel（扩展名骗人），但里面没找到 <table>。"
        )
    meta["sheet"] = path.name
    meta["source"] = "html(伪装成 xls)"
    meta["dimensions"] = f"{len(rows)} 行"
    width = max(len(r) for r in rows)
    for r in rows:
        r.extend([""] * (width - len(r)))
    return rows, meta


def profile(path: Path, sheet: str | None = None, max_scan: int = 200) -> TableProfile:
    grid, meta = read_grid(path, sheet)
    p = TableProfile(path=str(path), sheet=meta["sheet"], dimensions=meta["dimensions"])
    p.merged_ranges = meta["merged"]
    p.hidden_rows = meta["hidden_rows"]
    p.hidden_cols = meta["hidden_cols"]

    if not grid:
        p.warnings.append("空表")
        return p

    width = max(len(r) for r in grid)
    for r in grid:
        if len(r) < width:
            r.extend([None] * (width - len(r)))

    title_rows, header_rows, data_start = detect_header(grid, max_scan=min(max_scan, 40))
    p.title_rows = title_rows
    p.header_rows = header_rows
    p.multi_level_header = len(header_rows) > 1
    names = _build_names(grid, header_rows, width, p.merged_ranges)

    # ── 扫数据区，挑出汇总行、脚注行 ──────────────────────────────
    data_rows: list[int] = []          # 1-based 行号
    sparse_rows: list[dict] = []       # 待判定：零星的是小计，连成片的是并排子表
    for r in range(data_start, len(grid) + 1):
        row = grid[r - 1]
        if all(_is_blank(v) for v in row):
            continue
        # 首个非空格子的文本
        first_text = ""
        for v in row:
            if not _is_blank(v):
                first_text = _norm(v)
                break
        nonempty, text, num = _row_signature(row)

        if _looks_like_footnote(first_text):
            p.footnote_rows.append({"row": r, "text": first_text[:80]})
            continue
        if _looks_like_summary(first_text):
            p.summary_rows.append({"row": r, "label": first_text[:40],
                                   "reason": "标签含汇总类词汇"})
            continue
        # 稀疏行（大部分列空着但有数字）——典型的分区小计。
        # 先只记下来，等扫完再判：小计行是零星散布的，如果它们连成一大片，
        # 那就不是小计，是这张表右半边比左半边短（并排放了好几个子表）。
        if num >= 1 and nonempty <= max(2, width // 3):
            sparse_rows.append({"row": r, "label": first_text[:40],
                                "reason": "大部分列为空但含数字，疑似小计行"})
            continue
        data_rows.append(r)

    # 连成片的稀疏行不是小计行
    longest = _longest_run([s["row"] for s in sparse_rows])
    total = len(data_rows) + len(sparse_rows)
    if sparse_rows and (longest >= 10 or (total and len(sparse_rows) / total > 0.3)):
        first = sparse_rows[0]["row"]
        p.warnings.append(
            f"第 {first} 行起有 {len(sparse_rows)} 行「只有左边几列有值」，最长连续 {longest} 行——"
            f"小计行不会连成这么大一片，多半是这张表并排放了好几个子表、右边那些比左边短。"
            f"先按空列把子表切开再分别体检，不要当成一张表算。"
        )
        data_rows.extend(s["row"] for s in sparse_rows)
        data_rows.sort()
    else:
        p.summary_rows.extend(sparse_rows)

    if not data_rows:
        p.warnings.append("没有识别出明细数据行，请人工确认表头位置")
        p.data_start_row = data_start
        return p

    p.data_start_row = data_rows[0]
    p.data_end_row = data_rows[-1]
    p.n_data_rows = len(data_rows)

    # ── 合并单元格落在数据区 ────────────────────────────────────────
    merged_cols: set[int] = set()      # 1-based，受数据区合并影响的列
    if HAS_OPENPYXL:
        for rng in p.merged_ranges:
            m = RANGE_RE.match(rng.strip().upper())
            if not m:
                continue
            c1, r1 = _col_to_idx(m.group(1)), int(m.group(2))
            c2, r2 = _col_to_idx(m.group(3)), int(m.group(4))
            if r2 >= p.data_start_row and r1 <= p.data_end_row:
                p.merged_in_data.append({
                    "range": rng,
                    "note": "数据区内的合并单元格：只有左上角有值，读进 pandas 后其余行变成空值。"
                            "分组标签列常见此形态，需要向下填充（ffill）。",
                })
                merged_cols.update(range(c1, c2 + 1))

    # ── 重复行 ──────────────────────────────────────────────────────
    seen: dict[tuple, int] = {}
    for r in data_rows:
        key = tuple(_norm(v) for v in grid[r - 1])
        if key in seen:
            p.duplicate_rows.append({"row": r, "same_as": seen[key],
                                     "preview": " | ".join(x for x in key if x)[:80]})
        else:
            seen[key] = r

    # ── 逐列体检 ────────────────────────────────────────────────────
    for c in range(width):
        col = ColumnProfile(index=c, letter=get_column_letter(c + 1),
                            name=names[c] if c < len(names) else get_column_letter(c + 1))
        raws = [grid[r - 1][c] for r in data_rows]
        col.n_total = len(raws)

        # 先判标识列——它决定这一列的整数要不要当日期序列号看。
        col.identifier_reason = sniff_identifier(col.name, raws)
        col.is_identifier = bool(col.identifier_reason)

        disguised = Counter()
        kinds = Counter()
        nums: list[float] = []
        marks = Counter()
        date_fmts = Counter()
        texts: list[str] = []
        serial_pending: list[tuple[Any, float]] = []   # 待定的日期序列号候选

        for raw in raws:
            if raw is None or (isinstance(raw, float) and math.isnan(raw)):
                col.n_missing += 1
                continue
            s = _norm(raw)
            if s == "":
                col.n_missing += 1
                continue
            if isinstance(raw, str) and _is_null_token(s):
                col.n_missing += 1
                disguised[s] += 1
                continue

            dfmt = sniff_date(raw, allow_serial=not col.is_identifier)
            if dfmt in AMBIGUOUS_DATE_LABELS:
                # 先记着，等看完整列再判——零星几个长得像日期的纯数字是编号或
                # 金额，不是日期。守卫在下面的 serial_pending 处理里。
                serial_pending.append((raw, dfmt))
                continue
            if dfmt:
                kinds["date"] += 1
                date_fmts[dfmt] += 1
                continue

            val, mk = parse_number(raw)
            if val is not None:
                kinds["number"] += 1
                nums.append(val)
                for m in mk:
                    marks[m] += 1
                if isinstance(raw, str):
                    kinds["_number_stored_as_text"] += 1
            else:
                kinds["text"] += 1
                texts.append(s)

        # ── 日期序列号的列级守卫 ──────────────────────────────────────
        # 只有当「像序列号的整数」占了这一列相当比例，才认为它真是日期列。
        # 否则那些整数是编号或金额，原样还给数值分支。
        if serial_pending:
            n_live = col.n_total - col.n_missing
            share = len(serial_pending) / n_live if n_live else 0
            if share >= SERIAL_DATE_MIN_SHARE:
                for _raw_v, lbl in serial_pending:
                    kinds["date"] += 1
                    date_fmts[lbl] += 1
            else:
                for raw_v, _lbl in serial_pending:
                    val, mk = parse_number(raw_v)
                    if val is None:
                        continue
                    kinds["number"] += 1
                    nums.append(val)
                    for m in mk:
                        marks[m] += 1
                    if isinstance(raw_v, str):
                        kinds["_number_stored_as_text"] += 1

        col.n_unique = len({_norm(v) for v in raws if not _is_blank(v)})
        col.missing_disguised = dict(disguised)
        col.samples = [_norm(v) for v in raws[:4] if not _is_blank(v)]

        # 判定真实类型
        live = col.n_total - col.n_missing
        if live == 0:
            col.inferred_kind = "全空"
        else:
            top = max((k for k in ("number", "date", "text") if kinds.get(k)),
                      key=lambda k: kinds[k], default="text")
            col.inferred_kind = {"number": "数值", "date": "日期", "text": "文本"}[top]
            # 标识列一律按文本对待：编号不是数量，它的均值、方差、离群点都没有含义，
            # 下游脚本靠 inferred_kind 挑度量列，这里不改的话它们会去算「科目编号的均值」。
            if col.is_identifier:
                col.inferred_kind = "文本"
            mixed = sum(1 for k in ("number", "date", "text") if kinds.get(k, 0) > 0)
            if mixed > 1:
                parts = ", ".join(f"{k}×{kinds[k]}" for k in ("number", "date", "text") if kinds.get(k))
                if col.is_identifier:
                    col.issues.append(
                        f"数字与字母数字混排（{parts}）——标识列本来就长这样"
                        f"（536365 与 C489449 是同一种东西）。整列按文本处理即可，"
                        f"不要为了「统一类型」去做转换")
                else:
                    col.issues.append(f"混合类型（{parts}）——同一列里不同行是不同东西，聚合前必须先统一")

        # 类型层面的问题
        stored_as_text = kinds.get("_number_stored_as_text", 0)
        if stored_as_text:
            col.declared_kind = "文本"
            col.issues.append(
                f"{stored_as_text}/{live} 个数字被存成了文本 → pandas 读进来是字符串，"
                f"直接 sum() 会变成字符串拼接或得 0"
            )
        elif col.inferred_kind == "数值":
            col.declared_kind = "数值"

        if col.is_identifier:
            col.issues.append(
                f"标识列（{col.identifier_reason}）——保持文本，不要转数值。"
                f"转了会丢前导零、把 join 键变成浮点，而它的均值和离群点没有含义"
            )
            # 标识列里那些「不是纯数字」的值往往不是脏数据，是业务语义：
            # 发票号里的 C 是取消单、A 是坏账调整，科目码里的字母是特殊类别。
            # 不做转换是对的，但**必须把它们抛出来给人看**——
            # 一列编号里混着的那些异形值，常常是整份数据里最贵的一层信息。
            odd = {}
            for raw in raws:
                if not isinstance(raw, str):
                    continue
                v = raw.strip()
                if not v or v.isdigit():
                    continue
                # 按「前导非数字部分」归类，一类留一个代表值
                key = re.sub(r"\d+", "#", v)[:12]
                if key not in odd:
                    odd[key] = v
                if len(odd) >= 8:
                    break
            if odd:
                n_odd = sum(1 for r in raws
                            if isinstance(r, str) and r.strip() and not r.strip().isdigit())
                col.issues.append(
                    f"其中 {n_odd} 个值不是纯数字，形态举例："
                    f"{sorted(odd.values())[:6]} —— **去看看它们是什么，不要放过。**"
                    f"编号列里的异形值通常是业务语义（取消单、调整单、特殊类别），"
                    f"不是脏数据"
                )

        for m, cnt in marks.items():
            col.issues.append(f"{cnt} 个值带「{m}」")

        if len(date_fmts) > 1:
            col.issues.append(
                "日期格式不统一：" + "、".join(f"{k}×{v}" for k, v in date_fmts.items())
                + " —— 混排的日期没法直接排序或做时间序列"
            )
        elif date_fmts:
            fmt = next(iter(date_fmts))
            if fmt == "Excel日期序列号":
                col.issues.append("日期以 Excel 序列号存储（1900 日期系统），需要转换才能读")
            elif "缺年份" in fmt:
                col.issues.append("日期缺年份，跨年数据会错序")

        if disguised:
            col.issues.append(
                "缺失值被写成了具体文字：" + "、".join(f"「{k}」×{v}" for k, v in disguised.items())
                + " —— pandas 不认得这些，会当成正常字符串"
            )

        if col.n_total and col.n_missing == col.n_total:
            col.issues.append("整列为空")
        elif (c + 1) in merged_cols and col.n_missing:
            # 关键区分：这一列的「空」不是缺失，是合并单元格的显示效果。
            # 当成缺失去 dropna 会把整片分组删掉；正确做法是向下填充恢复标签。
            col.issues.append(
                f"{col.n_missing}/{col.n_total} 个空值来自合并单元格，不是真缺失 —— "
                f"要 ffill 向下填充恢复分组标签。若按缺失处理（dropna），"
                f"会把同组的其余行整片丢掉"
            )
        elif col.n_total and col.n_missing / col.n_total > 0.5:
            col.issues.append(f"缺失率 {col.n_missing / col.n_total:.0%}，超过一半")

        if live > 1 and col.n_unique == 1:
            col.issues.append("常量列（所有行同一个值），对分析没有信息量")

        # 前后空白/全角空格
        raw_strs = [v for v in raws if isinstance(v, str)]
        dirty_ws = sum(1 for v in raw_strs if v != v.strip() or FULLWIDTH_SPACE in v)
        if dirty_ws:
            col.issues.append(
                f"{dirty_ws} 个值带首尾空格或全角空格 —— 分组统计时「张伟」和「张伟 」会被算成两个人"
            )

        # 标识列不做五数概括。上面刚刚写完「它的均值和离群点没有含义」，
        # 紧接着又印一遍 min/Q1/中位数/均值，是自己打自己的脸——
        # 而且「税号的均值是 9.1×10^17」这种数印在报告里，读者会当它是个数量。
        if nums and not col.is_identifier:
            col.stats = five_number(nums)

        p.columns.append(col)

    # ── 读取建议 ────────────────────────────────────────────────────
    skip = sorted(set(p.title_rows) | {r["row"] for r in p.summary_rows}
                  | {r["row"] for r in p.footnote_rows})
    reader = "pd.read_csv" if meta.get("source") == "csv" else "pd.read_excel"
    p.read_hint = {
        "reader": reader,
        "header_rows_1based": p.header_rows,
        "pandas_header_arg": ([r - 1 for r in p.header_rows]
                              if p.multi_level_header else p.header_rows[0] - 1),
        "data_range_1based": [p.data_start_row, p.data_end_row],
        "rows_to_exclude_1based": skip,
        "note": "把上面这些行排除掉再读。汇总行混进明细是「算出来的数比真值大一倍还不报错」的头号原因。",
    }

    # ── 全表级警告 ──────────────────────────────────────────────────
    for w in _HEADER_WARNINGS:
        p.warnings.append(w)
    if p.title_rows:
        p.warnings.append(
            f"表头不在第 1 行：第 {_fmt_rows(p.title_rows)} 行是标题/说明，"
            f"真表头在第 {_fmt_rows(p.header_rows)} 行。"
            f"直接 pd.read_excel() 会把标题当成列名。"
        )
    if p.multi_level_header:
        p.warnings.append(
            f"两级及以上表头（第 {_fmt_rows(p.header_rows)} 行）。"
            f"用 header={p.read_hint['pandas_header_arg']} 读，否则下级表头会掉进数据区。"
        )
    if p.summary_rows:
        rows_s = _fmt_rows([r["row"] for r in p.summary_rows])
        p.warnings.append(
            f"发现 {len(p.summary_rows)} 个汇总/小计行（第 {rows_s} 行）混在数据区。"
            f"不排除的话，求和会把总计再加一遍。"
        )
    if p.duplicate_rows:
        p.warnings.append(
            f"{len(p.duplicate_rows)} 行与前面完全重复（第 "
            f"{_fmt_rows([r['row'] for r in p.duplicate_rows])} 行）。"
            f"先确认是真实重复业务还是粘贴事故。"
        )
    if p.merged_in_data:
        p.warnings.append(
            f"{len(p.merged_in_data)} 处合并单元格落在数据区。"
            f"pandas 只会保留左上角的值，其余行变空 —— 分组统计会整片丢数据。"
        )
    if p.hidden_rows or p.hidden_cols:
        p.warnings.append(
            f"存在隐藏行/列（行 {p.hidden_rows or '无'}，列 {p.hidden_cols or '无'}）。"
            f"隐藏往往意味着「有意排除」，纳入统计前先问清楚。"
        )
    return p


# ── 输出 ──────────────────────────────────────────────────────────────
def render(p: TableProfile) -> str:
    L: list[str] = []
    add = L.append
    bar = "─" * 68

    add(bar)
    add(f"表结构体检：{Path(p.path).name}  [{p.sheet}]")
    add(bar)
    add(f"原始范围 {p.dimensions}   识别出明细数据 {p.n_data_rows} 行 "
        f"(第 {p.data_start_row}-{p.data_end_row} 行) × {len(p.columns)} 列")

    if p.warnings:
        add("")
        add("【必须先处理】")
        for i, w in enumerate(p.warnings, 1):
            add(f"  {i}. {w}")

    add("")
    add("【怎么读这张表】")
    h = p.read_hint
    add(f"  pandas:  {h.get('reader', 'pd.read_excel')}(f, header={h.get('pandas_header_arg')})")
    if h.get("rows_to_exclude_1based"):
        add(f"  排除行(1-based):  {h['rows_to_exclude_1based']}")
    add(f"  明细区间(1-based):  {h.get('data_range_1based')}")

    if p.summary_rows:
        add("")
        add("【汇总行 / 小计行】不要计入明细统计")
        for r in p.summary_rows:
            add(f"  第{r['row']:>3}行  {r['label'] or '(无标签)':<16} {r['reason']}")

    if p.footnote_rows:
        add("")
        add("【脚注行】")
        for r in p.footnote_rows:
            add(f"  第{r['row']:>3}行  {r['text']}")

    if p.merged_in_data:
        add("")
        add("【数据区的合并单元格】")
        for m in p.merged_in_data:
            add(f"  {m['range']}")
        add(f"  → {p.merged_in_data[0]['note']}")

    if p.duplicate_rows:
        add("")
        add("【重复行】")
        for d in p.duplicate_rows:
            add(f"  第{d['row']}行 ≡ 第{d['same_as']}行   {d['preview']}")

    add("")
    add("【逐列】")
    for c in p.columns:
        miss = f"{c.n_missing}/{c.n_total}" if c.n_total else "-"
        add(f"  {c.letter}  {c.name}")
        add(f"      类型={c.inferred_kind}  缺失={miss}  唯一值={c.n_unique}"
            + (f"  样例={c.samples[:3]}" if c.samples else ""))
        for issue in c.issues:
            add(f"      ⚠ {issue}")
        if c.stats:
            s = c.stats
            add(f"      五数概括  min={s['min']:,.2f}  Q1={s['q1']:,.2f}  "
                f"中位数={s['median']:,.2f}  Q3={s['q3']:,.2f}  max={s['max']:,.2f}")
            add(f"                均值={s['mean']:,.2f}  （均值仅供参考，偏态数据看中位数）")
            if s.get("outlier_count"):
                ex = ", ".join(f"{x:,.0f}" for x in s["outlier_examples"])
                add(f"      离群点  {s['outlier_count']} 个（IQR 1.5 倍法则）：{ex}")
            if s.get("distribution_note"):
                add(f"      分布  {s['distribution_note']}")
    add("")
    add(bar)
    add("体检完毕。在读到「必须先处理」清空之前，不要开始算任何业务数字。")
    add(bar)
    return "\n".join(L)


def main() -> None:
    ap = argparse.ArgumentParser(description="表结构体检：算数字之前先看清这张表")
    ap.add_argument("file")
    ap.add_argument("--sheet", default=None, help="工作表名，默认活动表")
    ap.add_argument("--json", action="store_true", help="输出 JSON 便于程序消费")
    ap.add_argument("--max-scan", type=int, default=200, help="表头侦测扫描行数上限")
    a = ap.parse_args()

    path = Path(a.file).expanduser()
    if not path.exists():
        sys.exit(f"文件不存在：{path}")

    p = profile(path, a.sheet, a.max_scan)
    if a.json:
        print(json.dumps(asdict(p), ensure_ascii=False, indent=2, default=str))
    else:
        print(render(p))


if __name__ == "__main__":
    main()
