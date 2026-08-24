#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""把洞察渲染成一份自包含的 HTML 报告。

分工
----
洞察由 agent 产出（那是它擅长的），本脚本只负责渲染和画图——
这样图表的诚实性约束、口径声明、边界声明这几件容易被省略的事，
就变成了结构性保证而不是靠自觉。

零外部资源
----------
无 CDN、无外部字体、无网络请求。图表是自己画的内联 SVG。
所以它在 file:// 下、在离线环境、在内网都能正常打开，
也能直接当附件发给别人。

用法
----
    python3 make_report.py --spec report.json --out report.html
    python3 make_report.py --template > report.json      # 拿一份模板改

spec 结构见 --template。风格由 agent 按内容判断后填进 spec，
不要拿去问用户——报告不是设计作品，选风格是你的活。

依赖：无。纯标准库。
"""

from __future__ import annotations

import argparse
import html
import json
import sys
from pathlib import Path
from typing import Any

# ── 风格库 ──────────────────────────────────────────────────────────
# 每种风格取自一类机构的设计语言。真正的区别不在配色，在结构：
# 咨询给每张图编号并把结论写进图题，投行把表格做到最高密度，
# 财经媒体用衬线和窄栏做叙事，科技公司靠留白和圆角。
#
# 选哪种是 agent 的活，不是用户的活（见 references/report-styles.md）。
# 都不合适就用 custom 自己配。
SANS = "-apple-system,'PingFang SC','Microsoft YaHei','Hiragino Sans GB',Helvetica,sans-serif"
SERIF = "'Songti SC','Source Han Serif SC','Noto Serif CJK SC',Georgia,'Times New Roman',serif"

STYLES = {
    "consulting": {
        "desc": "咨询公司。深海军蓝，每张图编号 Exhibit N，图题即结论。战略分析、框架评估、董事会材料",
        "bg": "#FFFFFF", "fg": "#051C2C", "muted": "#5A6E7C",
        "rule": "#CBD5DD", "card": "#E7EEF2", "accent": "#2251FF",
        "pos": "#00806A", "neg": "#C1372B",
        "h_font": SANS, "b_font": SANS,
        "exhibit": True,
        "extra_css": """
h1{letter-spacing:-.02em;font-weight:700;border-bottom:3px solid %(accent)s;
padding-bottom:14px;display:inline-block}
h2{border-bottom:0;font-size:19px;color:%(fg)s;
padding-left:12px;border-left:4px solid %(accent)s}
.chart-title{font-size:16px;line-height:1.5}
.exhibit{font-size:11px;letter-spacing:.11em;text-transform:uppercase;
color:%(accent)s;font-weight:700;margin-bottom:5px}
.lede{border-left-width:4px;background:%(card)s}
.kpi{border-top:3px solid %(accent)s;border-radius:0}
th{text-transform:uppercase;letter-spacing:.05em;font-size:11.5px}
""",
    },
    "bank": {
        "desc": "投行研究报告。深蓝、高信息密度、等宽数字、细密分隔线。财务建模、估值、对账、尽调",
        "bg": "#FFFFFF", "fg": "#0B1F3A", "muted": "#63748A",
        "rule": "#CED5E0", "card": "#E8ECF3", "accent": "#0B4F8C",
        "pos": "#0F6B3D", "neg": "#A01B22",
        "h_font": SANS, "b_font": SANS,
        "extra_css": """
body{font-size:15px;line-height:1.65}
.wrap{max-width:1180px}
h1{font-size:26px;font-weight:700}
h2{font-size:17px;margin:38px 0 10px;text-transform:none;
border-bottom:2px solid %(fg)s;padding-bottom:6px}
table{font-size:13px}
th,td{padding:6px 10px}
tbody tr:nth-child(even){background:%(card)s}
.kpi{border-radius:3px;padding:12px 14px}
.kpi .v{font-size:22px}
.caliber{border-radius:3px;font-size:13px}
.lede{border-radius:0;font-size:15.5px}
""",
    },
    "editorial": {
        "desc": "财经媒体。三文鱼粉底、衬线标题、窄栏叙事。行业观察、深度解读、有观点的报告",
        "bg": "#FFF1E5", "fg": "#33302E", "muted": "#66605C",
        "rule": "#D9C4AC", "card": "#FFF9F1", "accent": "#0F5499",
        "pos": "#0D7680", "neg": "#CC0000",
        "h_font": SERIF, "b_font": SANS,
        "extra_css": """
.wrap{max-width:900px}
h1{font-size:34px;line-height:1.28}
h2{font-family:%(h_font)s;border-bottom:0;font-size:22px;
margin:48px 0 12px}
p{font-size:16.5px;line-height:1.8}
.lede{background:transparent;border-left:3px solid %(accent)s;
font-family:%(h_font)s;font-size:18px}
.kpi{background:%(card)s;border-color:%(rule)s}
.chart{border-top:1px solid %(rule)s;padding-top:16px}
""",
    },
    "magazine": {
        "desc": "杂志式。白底红标块、紧凑排版、editorial 标题。观点报告、行业洞察、有立场的分析",
        "bg": "#FFFFFF", "fg": "#121212", "muted": "#6E6E6E",
        "rule": "#D5D5D5", "card": "#ECECEC", "accent": "#E3120B",
        "pos": "#0B7A3B", "neg": "#E3120B",
        "h_font": SANS, "b_font": SANS,
        "extra_css": """
.wrap{max-width:880px}
h1{font-size:29px;line-height:1.3;font-weight:700}
h1::before{content:"";display:block;width:52px;height:7px;
background:%(accent)s;margin-bottom:15px}
h2{border-bottom:0;font-size:18px;margin:44px 0 10px;
padding-left:0}
h2::before{content:"";display:inline-block;width:14px;height:3px;
background:%(accent)s;vertical-align:middle;margin-right:9px}
.chart-title::before{content:"";display:inline-block;width:9px;height:9px;
background:%(accent)s;margin-right:8px}
body{font-size:15.5px}
.lede{background:%(card)s;border-left:3px solid %(accent)s}
""",
    },
    "product": {
        "desc": "科技公司。大留白、圆角卡片、克制的紫蓝强调。产品复盘、增长分析、内部评审",
        "bg": "#FFFFFF", "fg": "#0A2540", "muted": "#66768C",
        "rule": "#CBD5E2", "card": "#E7EFF7", "accent": "#635BFF",
        "pos": "#09825D", "neg": "#CD3D64",
        "h_font": SANS, "b_font": SANS,
        "extra_css": """
.wrap{max-width:1000px;padding-top:72px}
h1{font-size:33px;letter-spacing:-.025em;font-weight:700}
h2{border-bottom:0;font-size:21px;margin:56px 0 12px;letter-spacing:-.015em}
.kpi{border-radius:12px;border-color:%(rule)s;background:%(card)s;padding:18px 20px}
.kpi .v{letter-spacing:-.03em}
.lede{border-radius:12px;border-left:0;background:%(card)s;padding:22px 24px}
.caliber{border-radius:12px}
.bounds{border-radius:12px}
.chart{background:%(card)s;border-radius:12px;padding:20px 22px}
table{font-size:14.5px}
""",
    },
    "minimal": {
        "desc": "极简。大量留白、无装饰。内部快看、单一结论、只想把数说清楚",
        "bg": "#FFFFFF", "fg": "#000000", "muted": "#757575",
        "rule": "#D5D5D5", "card": "#EDEDED", "accent": "#000000",
        "pos": "#1F6F3F", "neg": "#95291F",
        "h_font": SANS, "b_font": SANS,
        "extra_css": """
.wrap{max-width:860px}
h2{border-bottom:0;font-size:18px;margin:48px 0 10px}
.lede{background:transparent;border-left:2px solid %(fg)s}
.kpi{border:0;background:transparent;padding:0 20px 0 0}
.kpis{gap:32px}
""",
    },
}
# 别名，向后兼容
STYLES["report"] = STYLES["editorial"]
STYLES["finance"] = STYLES["bank"]

# 色盲友好色板（Wong 系）：避开红绿对立，约 8% 男性有红绿色觉障碍
SERIES_COLORS = ["#0173B2", "#DE8F05", "#029E73", "#CC78BC", "#CA9161", "#56B4E9"]

E = html.escape


def _nice_num(x: float, round_it: bool) -> float:
    """把一个数收成「好看的数」：1 / 2 / 2.5 / 5 的 10 的幂倍。"""
    import math
    if x <= 0:
        return 1.0
    exp = math.floor(math.log10(x))
    f = x / (10 ** exp)
    if round_it:
        nf = 1 if f < 1.5 else (2 if f < 3 else (5 if f < 7 else 10))
    else:
        nf = 1 if f <= 1 else (2 if f <= 2 else (5 if f <= 5 else 10))
    return nf * (10 ** exp)


def nice_scale(lo: float, hi: float, ticks: int = 4) -> tuple[float, float, float]:
    """算出整齐的坐标轴刻度。

    139.68 / 115.50 / 91.32 这种从数据直接算出来的碎刻度，
    读者要多花一秒去解析每个数字。整刻度（100/120/140）是免费的可读性。

    注意：这里只把刻度收整，不强制从 0 —— 折线编码的是位置和斜率，
    强行归零会把真实波动压平，那是另一种误导。
    """
    import math
    if hi <= lo:
        hi, lo = lo + 1, lo - 1
    # 直接由真实跨度算 step。先把 span 收整一次再除，等于向上取整叠加两次，
    # 实测能把轴撑到数据跨度的 2 倍以上，数据被压成中间一条窄带。
    step = _nice_num((hi - lo) / max(ticks - 1, 1), True)
    return math.floor(lo / step) * step, math.ceil(hi / step) * step, step


def _fmt(v: float) -> str:
    if v is None:
        return "—"
    a = abs(v)
    if a >= 1e8:
        return f"{v/1e8:,.2f}亿"
    if a >= 1e4:
        return f"{v/1e4:,.1f}万"
    if a == int(a):
        return f"{int(v):,}"
    return f"{v:,.2f}"


def _decimals(values: list) -> int:
    """一组数里最多用到几位小数（上限 2）。"""
    d = 0
    for v in values:
        if not isinstance(v, (int, float)) or isinstance(v, bool):
            continue
        if v != v or v in (float("inf"), float("-inf")):   # NaN / inf
            continue
        for k in (0, 1, 2):
            if round(float(v), k) == float(v):
                d = max(d, k)
                break
        else:
            d = 2
    return d


def fmt_column(values: list) -> "callable":
    """给一整列返回统一的格式化函数。

    逐个格式化会让同一列里出现 `0.90` 和 `0`、`313.80` 和 `75` 并排——
    读者要在每一行重新判断小数点在哪。同列同格式是最基本的表格素养。

    表格里不做万/亿缩写：表是拿来核数的，缩写会丢精度。
    """
    d = _decimals(values)
    def f(v):
        if v is None or not isinstance(v, (int, float)) or isinstance(v, bool):
            return "—" if v is None else str(v)
        return f"{v:,.{d}f}"
    return f


def fmt_axis(step: float) -> "callable":
    """坐标轴刻度的统一格式化。

    刻度由 step 决定小数位，否则会出现 `4.50 / 4 / 3.50 / 3` 这种
    同一根轴上格式跳来跳去的情况。
    """
    d = _decimals([step])
    def f(v):
        a = abs(v)
        if a >= 1e8:
            return f"{v/1e8:,.2f}亿"
        if a >= 1e4:
            return f"{v/1e4:,.1f}万"
        return f"{v:,.{d}f}"
    return f


# ── SVG 图表：自己画，不依赖任何库 ──────────────────────────────────
def _label_w(s: str, size: float = 13.0) -> float:
    """文本渲染宽度估算。中文字符宽度约等于字号，西文数字约 0.55 倍。
    所有留白都要按它算，不能写死：写死 130 时「豆包手搓」这种标签
    会被 viewBox 边界裁掉首字，而且裁掉的部分不会有任何报错；
    按 len() 算则中文必然溢出——中文一个字符占两倍宽。"""
    wide = sum(1 for c in s if ord(c) > 0x2E80)
    return wide * size + (len(s) - wide) * size * 0.55


def svg_bar(labels: list[str], values: list[float], st: dict,
            highlight: int | None = None) -> str:
    """横向条形图。长度 + 共同基线是感知精度最高的编码方式
    （Cleveland & McGill 1984），所以这是默认图表。"""
    if not values:
        return ""
    n = len(values)
    row_h, gap, pad_r, pad_t = 30, 8, 90, 8

    LABEL_MAX = 190.0
    labels = [str(x) for x in labels]
    shown = []
    for lab in labels:
        if _label_w(lab) <= LABEL_MAX:
            shown.append(lab)
            continue
        cut = lab                      # 超长的截断，但保留完整值给 <title>
        while cut and _label_w(cut + "…") > LABEL_MAX:
            cut = cut[:-1]
        shown.append(cut + "…")
    pad_l = min(max(max((_label_w(s) for s in shown), default=0) + 16, 90.0), LABEL_MAX + 16)
    w = pad_l + 420 + pad_r            # 条形区至少留 420，总宽随标签伸缩
    h = pad_t * 2 + n * row_h + (n - 1) * gap
    plot_w = w - pad_l - pad_r
    # 基线从 0：条形编码长度，截断即失真
    vmax = max(max(values), 0)
    vmin = min(min(values), 0)
    span = (vmax - vmin) or 1
    zero_x = pad_l + (0 - vmin) / span * plot_w

    barf = fmt_column([v for v in values if v is not None])
    parts = [f'<svg viewBox="0 0 {w} {h}" width="100%" '
             f'preserveAspectRatio="xMidYMid meet" role="img">']
    for i, (lab, disp, val) in enumerate(zip(labels, shown, values)):
        y = pad_t + i * (row_h + gap)
        x1 = pad_l + (min(val, 0) - vmin) / span * plot_w
        x2 = pad_l + (max(val, 0) - vmin) / span * plot_w
        bw = max(x2 - x1, 1.5)
        color = st["accent"] if (highlight is not None and i == highlight) \
            else SERIES_COLORS[0]
        op = "1" if (highlight is None or i == highlight) else "0.45"
        parts.append(
            f'<rect x="{x1:.1f}" y="{y}" width="{bw:.1f}" height="{row_h}" '
            f'rx="2" fill="{color}" opacity="{op}"/>')
        title = f'<title>{E(lab)}</title>' if disp != lab else ""
        parts.append(
            f'<text x="{pad_l - 10:.0f}" y="{y + row_h*0.68:.0f}" text-anchor="end" '
            f'font-size="13" fill="{st["fg"]}">{title}{E(disp)}</text>')
        # 直接标注数值，不用图例也不用坐标轴刻度
        parts.append(
            f'<text x="{x2 + 8:.1f}" y="{y + row_h*0.68:.0f}" font-size="13" '
            f'font-variant-numeric="tabular-nums" fill="{st["muted"]}">{barf(val)}</text>')
    if vmin < 0:
        parts.append(f'<line x1="{zero_x:.1f}" y1="{pad_t-4}" x2="{zero_x:.1f}" '
                     f'y2="{h-pad_t+4}" stroke="{st["rule"]}" stroke-width="1"/>')
    parts.append("</svg>")
    return "".join(parts)


def svg_line(labels: list[str], series: list[dict], st: dict,
             incomplete_last: bool = False) -> str:
    """折线图。不强制 Y 轴从 0——折线编码位置和斜率，强行归零会压平真实波动。"""
    if not series or not series[0].get("values"):
        return ""
    w, h = 720, 300
    pad_l, pad_t, pad_b = 56, 16, 34
    # 右边距装线末系列名，按实际渲染宽度算。写死 80 时
    # 「累计同比（官方）」这类中文系列名必然溢出 viewBox（06 号压测实锤）。
    names = [str(s.get("name", "") or "") for s in series]
    pad_r = min(max([_label_w(nm, 12) + 16 for nm in names] + [56.0]), 220.0)
    pw, ph = w - pad_l - pad_r, h - pad_t - pad_b
    allv = [v for s in series for v in s["values"] if v is not None]
    if not allv:
        return ""
    dlo, dhi = min(allv), max(allv)
    pad = (dhi - dlo) * 0.12 if dhi > dlo else max(abs(dhi) * 0.1, 1)
    lo, hi, step = nice_scale(dlo - pad, dhi + pad, ticks=4)
    # 不强制从 0（见 nice_scale 的说明），但数据本身全非负时，轴下界不该跌破 0：
    # 「阅读人数 -2万」这样的刻度是假的，且下方那段空白不携带任何波动信息。
    if dlo >= 0 and lo < 0:
        lo = 0.0
    n = len(labels)

    def X(i: int) -> float:
        return pad_l + (i / max(n - 1, 1)) * pw

    def Y(v: float) -> float:
        return pad_t + (1 - (v - lo) / (hi - lo)) * ph

    parts = [f'<svg viewBox="0 0 {w} {h}" width="100%" '
             f'preserveAspectRatio="xMidYMid meet" role="img">']
    # 极淡的水平参考线，落在整刻度上
    axf = fmt_axis(step)
    val = lo
    while val <= hi + step * 0.001:
        y = Y(val)
        parts.append(f'<line x1="{pad_l}" y1="{y:.1f}" x2="{pad_l+pw}" y2="{y:.1f}" '
                     f'stroke="{st["rule"]}" stroke-width="1"/>')
        parts.append(f'<text x="{pad_l-8}" y="{y+4:.1f}" text-anchor="end" '
                     f'font-size="11" fill="{st["muted"]}" '
                     f'font-variant-numeric="tabular-nums">{axf(val)}</text>')
        val += step

    end_labels: list[tuple[float, float, str, str]] = []
    for si, s in enumerate(series):
        color = SERIES_COLORS[si % len(SERIES_COLORS)]
        vals = s["values"]
        pts = [(X(i), Y(v)) for i, v in enumerate(vals) if v is not None]
        if not pts:
            continue
        # 不完整的最后一期用虚线，而不是画成一次真实的下跌
        if incomplete_last and len(pts) > 1:
            solid = pts[:-1]
            parts.append('<polyline points="'
                         + " ".join(f"{x:.1f},{y:.1f}" for x, y in solid)
                         + f'" fill="none" stroke="{color}" stroke-width="2.2" '
                           f'stroke-linejoin="round"/>')
            (x1, y1), (x2, y2) = pts[-2], pts[-1]
            parts.append(f'<line x1="{x1:.1f}" y1="{y1:.1f}" x2="{x2:.1f}" y2="{y2:.1f}" '
                         f'stroke="{color}" stroke-width="2.2" stroke-dasharray="5 4"/>')
        else:
            parts.append('<polyline points="'
                         + " ".join(f"{x:.1f},{y:.1f}" for x, y in pts)
                         + f'" fill="none" stroke="{color}" stroke-width="2.2" '
                           f'stroke-linejoin="round"/>')
        for x, y in pts:
            parts.append(f'<circle cx="{x:.1f}" cy="{y:.1f}" r="3" fill="{color}"/>')
        # 直接在线末标注系列名，不做图例——但先收集，出循环统一排布
        end_labels.append((*pts[-1], color, str(s.get("name", "") or "")))

    # 线末标签的碰撞规避：两条线终点接近时标签会叠死（07 号压测实锤）。
    # 按 y 排序后强制拉开最小行距，整体夹回绘图区内。
    if end_labels:
        min_gap = 15.0
        order = sorted(range(len(end_labels)), key=lambda i: end_labels[i][1])
        ys = [end_labels[i][1] + 4 for i in order]
        for k in range(1, len(ys)):
            if ys[k] - ys[k - 1] < min_gap:
                ys[k] = ys[k - 1] + min_gap
        over = ys[-1] - (h - pad_b - 4)
        if over > 0:
            ys = [y - over for y in ys]
        for k, i in enumerate(order):
            lx, _ly, color, name = end_labels[i]
            parts.append(f'<text x="{lx+9:.1f}" y="{max(ys[k], pad_t + 10):.1f}" '
                         f'font-size="12" fill="{color}">{E(name)}</text>')

    step = max(1, n // 8)
    for i, lab in enumerate(labels):
        if i % step == 0 or i == n - 1:
            parts.append(f'<text x="{X(i):.1f}" y="{h-12}" text-anchor="middle" '
                         f'font-size="11" fill="{st["muted"]}">{E(str(lab)[:10])}</text>')
    parts.append("</svg>")
    return "".join(parts)


def _chart_svg(ch: dict, st: dict) -> str:
    """图表的 SVG 本体。html/deck 直接内联；docx 拿它渲成 PNG 内嵌。"""
    kind = ch.get("type", "bar")
    labels = [str(x) for x in ch.get("labels", [])]
    if kind == "line":
        series = ch.get("series") or [{"name": ch.get("name", ""),
                                       "values": ch.get("values", [])}]
        return svg_line(labels, series, st, ch.get("incomplete_last", False))
    vals = ch.get("values", [])
    pairs = list(zip(labels, vals))
    if ch.get("sort", True):
        pairs.sort(key=lambda t: (t[1] is None, -(t[1] or 0)))
    hl = ch.get("highlight")
    hi = None
    if hl is not None:
        for i, (lab, _v) in enumerate(pairs):
            if lab == str(hl):
                hi = i
                break
    return svg_bar([p[0] for p in pairs], [p[1] for p in pairs], st, hi)


def render_chart(ch: dict, st: dict, exhibit_no: int | None = None) -> str:
    body = _chart_svg(ch, st)
    if not body:
        return ""
    cap = ch.get("caption", "")
    note = ch.get("note", "")
    out = ['<figure class="chart">']
    if exhibit_no is not None:
        out.append(f'<div class="exhibit">Exhibit {exhibit_no}</div>')
    if cap:
        out.append(f'<figcaption class="chart-title">{E(cap)}</figcaption>')
    out.append(f'<div class="chart-body">{body}</div>')
    if note:
        out.append(f'<p class="chart-note">{E(note)}</p>')
    out.append("</figure>")
    return "".join(out)


def render_table(tb: dict, st: dict) -> str:
    cols = tb.get("columns", [])
    rows = tb.get("rows", [])
    if not cols or not rows:
        return ""
    out = ['<div class="table-wrap"><table>']
    if tb.get("caption"):
        out.append(f'<caption>{E(tb["caption"])}</caption>')
    out.append("<thead><tr>" + "".join(f"<th>{E(str(c))}</th>" for c in cols)
               + "</tr></thead><tbody>")
    # 每列独立算一次格式：同列同小数位，读者不用逐行重新找小数点
    ncol = max((len(r) for r in rows), default=0)
    colfmt = []
    for j in range(ncol):
        col = [r[j] for r in rows if j < len(r)]
        nums = [v for v in col if isinstance(v, (int, float)) and not isinstance(v, bool)]
        colfmt.append(fmt_column(nums) if nums else None)
    for r in rows:
        cells = []
        for j, v in enumerate(r):
            num = isinstance(v, (int, float)) and not isinstance(v, bool)
            txt = colfmt[j](v) if (num and j < len(colfmt) and colfmt[j]) else E(str(v))
            cells.append(f'<td class="{"num" if num else ""}">{txt}</td>')
        out.append("<tr>" + "".join(cells) + "</tr>")
    out.append("</tbody></table></div>")
    return "".join(out)


CSS = """
*{box-sizing:border-box}
body{margin:0;background:%(bg)s;color:%(fg)s;font-family:%(b_font)s;
line-height:1.75;font-size:16px;-webkit-font-smoothing:antialiased}
.wrap{max-width:1040px;margin:0 auto;padding:56px 32px 96px}
h1{font-family:%(h_font)s;font-size:31px;line-height:1.35;margin:0 0 10px;
font-weight:700;letter-spacing:-.01em}
h2{font-family:%(h_font)s;font-size:20px;margin:52px 0 14px;font-weight:700;
padding-bottom:9px;border-bottom:1px solid %(rule)s}
h3{font-size:16px;margin:26px 0 8px;font-weight:600}
p{margin:0 0 15px}
.sub{color:%(muted)s;font-size:14px;margin-bottom:30px}
.lede{font-size:17px;padding:20px 22px;background:%(card)s;
border-left:3px solid %(accent)s;border-radius:0 6px 6px 0;margin:0 0 32px}
.lede ul{margin:0;padding-left:19px}
.lede li{margin:7px 0}
.kpis{display:flex;flex-wrap:wrap;gap:12px;margin:0 0 30px}
.kpi{flex:1 1 150px;background:%(card)s;border:1px solid %(rule)s;
border-radius:8px;padding:15px 17px}
.kpi .k{font-size:12px;color:%(muted)s;margin-bottom:5px}
.kpi .v{font-size:25px;font-weight:700;font-variant-numeric:tabular-nums;
letter-spacing:-.02em}
.kpi .d{font-size:12px;color:%(muted)s;margin-top:3px}
.pos{color:%(pos)s}.neg{color:%(neg)s}
.chart{margin:22px 0 26px}
.chart-title{font-size:15px;font-weight:600;margin-bottom:10px}
.chart-body{overflow-x:auto}
.chart-note{font-size:12.5px;color:%(muted)s;margin:8px 0 0}
.table-wrap{overflow-x:auto;margin:18px 0 24px}
table{border-collapse:collapse;width:100%%;font-size:14px}
caption{text-align:left;font-size:13px;color:%(muted)s;padding-bottom:8px}
th,td{padding:8px 11px;border-bottom:1px solid %(rule)s;text-align:left}
th{font-weight:600;font-size:13px;color:%(muted)s;
border-bottom:1.5px solid %(rule)s;white-space:nowrap}
td:first-child{white-space:nowrap}
td.num{text-align:right;font-variant-numeric:tabular-nums}
tbody tr:hover{background:%(card)s}
.caliber{background:%(card)s;border:1px solid %(rule)s;border-radius:8px;
padding:17px 20px;margin:0 0 30px;font-size:14px}
.caliber dl{margin:0;display:grid;grid-template-columns:auto 1fr;gap:5px 16px}
.caliber dt{color:%(muted)s;white-space:nowrap}
.caliber dd{margin:0}
.bounds{border:1px dashed %(rule)s;border-radius:8px;padding:17px 20px;
margin:40px 0 0;font-size:14px;color:%(muted)s}
.bounds h2{border:0;margin:0 0 9px;font-size:15px;padding:0;color:%(fg)s}
.bounds ul{margin:0;padding-left:19px}
.bounds li{margin:5px 0}
footer{margin-top:44px;padding-top:16px;border-top:1px solid %(rule)s;
font-size:12px;color:%(muted)s}
@media(max-width:600px){.wrap{padding:32px 18px 64px}h1{font-size:25px}}
@media print{body{background:#fff}.wrap{padding:0}
.chart,.table-wrap,.kpi{break-inside:avoid}}
"""


def resolve_style(spec: dict) -> dict:
    """定出这份报告用什么风格。

    优先级：spec.custom_style（agent 自己配的）> spec.style（内置名）> consulting。
    内置的都不合适时，agent 应该直接给 custom_style，不要将就一个不对的。
    """
    base = dict(STYLES.get(spec.get("style", "consulting"), STYLES["consulting"]))
    custom = spec.get("custom_style")
    if isinstance(custom, dict):
        # 只允许覆盖表现层字段，防止 spec 注入结构性内容
        allowed = {"bg", "fg", "muted", "rule", "card", "accent", "pos", "neg",
                   "h_font", "b_font", "extra_css", "exhibit", "desc"}
        base.update({k: v for k, v in custom.items() if k in allowed})
    base.setdefault("extra_css", "")
    base.setdefault("exhibit", False)
    return base


def build_html(spec: dict) -> str:
    st = resolve_style(spec)
    exhibit_on = bool(st.get("exhibit")) or bool(spec.get("exhibit"))
    exhibit_n = 0

    body: list[str] = []
    body.append(f'<h1>{E(spec.get("title", "数据分析报告"))}</h1>')
    if spec.get("subtitle"):
        body.append(f'<p class="sub">{E(spec["subtitle"])}</p>')

    kpis = spec.get("kpis") or []
    if kpis:
        cards = []
        for k in kpis:
            val = k.get("value")
            vs = _fmt(val) if isinstance(val, (int, float)) else E(str(val))
            cls = ""
            d = k.get("delta")
            if isinstance(d, (int, float)):
                cls = "pos" if d > 0 else ("neg" if d < 0 else "")
            dtxt = ""
            if k.get("note"):
                dtxt = f'<div class="d {cls}">{E(str(k["note"]))}</div>'
            cards.append(f'<div class="kpi"><div class="k">{E(str(k.get("label","")))}'
                         f'</div><div class="v">{vs}</div>{dtxt}</div>')
        body.append('<div class="kpis">' + "".join(cards) + "</div>")

    summary = spec.get("summary") or []
    if summary:
        body.append('<div class="lede"><ul>'
                    + "".join(f"<li>{E(s)}</li>" for s in summary) + "</ul></div>")

    # 口径声明。结构性地放在正文之前——说不清口径的数字不能被检验。
    cal = spec.get("caliber") or {}
    if cal:
        items = "".join(f"<dt>{E(str(k))}</dt><dd>{E(str(v))}</dd>"
                        for k, v in cal.items())
        body.append(f'<div class="caliber"><dl>{items}</dl></div>')

    for sec in spec.get("sections", []):
        if sec.get("heading"):
            body.append(f'<h2>{E(sec["heading"])}</h2>')
        for para in (sec.get("body") or "").split("\n\n"):
            if para.strip():
                body.append(f"<p>{E(para.strip())}</p>")
        for ch in (sec.get("charts") or ([sec["chart"]] if sec.get("chart") else [])):
            if exhibit_on:
                exhibit_n += 1
            body.append(render_chart(ch, st, exhibit_n if exhibit_on else None))
        for tb in (sec.get("tables") or ([sec["table"]] if sec.get("table") else [])):
            body.append(render_table(tb, st))
        for sub in sec.get("subsections", []):
            if sub.get("heading"):
                body.append(f'<h3>{E(sub["heading"])}</h3>')
            for para in (sub.get("body") or "").split("\n\n"):
                if para.strip():
                    body.append(f"<p>{E(para.strip())}</p>")

    # 边界声明。固定结构，不由 agent 决定要不要写。
    bounds = spec.get("boundaries") or []
    if bounds:
        body.append('<div class="bounds"><h2>这份分析的边界</h2><ul>'
                    + "".join(f"<li>{E(b)}</li>" for b in bounds) + "</ul></div>")

    if spec.get("footer"):
        body.append(f'<footer>{E(spec["footer"])}</footer>')

    return f"""<!doctype html>
<html lang="zh-CN"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{E(spec.get('title', '数据分析报告'))}</title>
<style>{CSS % st}{(st.get('extra_css') or '') % st}</style></head>
<body><div class="wrap">{''.join(body)}</div></body></html>"""


# ── 形态二：HTML 幻灯片 ─────────────────────────────────────────────
DECK_CSS = """
*{box-sizing:border-box}
html,body{margin:0;padding:0;background:#1B1D21;height:100%%;
font-family:%(b_font)s;-webkit-font-smoothing:antialiased}
.slide{width:100vw;height:100vh;display:none;padding:5.5vh 7vw;
background:%(bg)s;color:%(fg)s;flex-direction:column;position:relative;
container-type:size}
.slide.on{display:flex}
.stitle{font-family:%(h_font)s;font-size:clamp(20px,2.9vw,36px);
line-height:1.3;font-weight:700;margin:0 0 2.6vh;letter-spacing:-.015em;
max-width:100%%;text-wrap:balance}
.slide:not(.cover) .stitle{border-bottom:2px solid %(fg)s;padding-bottom:1.8vh}
.sbody.one .scol{flex:1}
.sbody.one .chart-body svg{max-height:64vh}
.sfoot{border-top:1px solid %(rule)s;margin-top:2.2vh;padding-top:1.5vh;
font-size:clamp(11px,1.15vw,15px);color:%(muted)s;line-height:1.7}
.sfoot p{margin:0 0 .6vh}
.skicker{font-size:clamp(10px,1.05vw,13px);letter-spacing:.14em;
text-transform:uppercase;color:%(accent)s;font-weight:700;margin-bottom:1.4vh}
.sbody{flex:1;display:flex;gap:3.2vw;min-height:0}
.scol{flex:1;min-width:0;display:flex;flex-direction:column;justify-content:center}
.sbody .scol:first-child{flex:1.5}
.scol svg{max-height:62vh}
.scol>div{flex:0 1 100%%;min-height:0;display:flex;flex-direction:column;
justify-content:center}
.chart-title{font-size:clamp(13px,1.35vw,19px);font-weight:600;
color:%(fg)s;margin:0 0 1.6vh}
.chart-body{flex:0 1 100%%;min-height:0;display:flex;align-items:center}
.chart-body svg{width:100%%;height:100%%;max-height:58vh}
.chart-note{font-size:clamp(11px,1.05vw,14px);color:%(muted)s;margin:1.4vh 0 0}
.spoints{list-style:none;margin:0;padding:0;font-size:clamp(13px,1.45vw,21px);
line-height:1.65}
.spoints li{margin:0 0 1.5vh;padding-left:1.1em;position:relative}
.spoints li::before{content:"";position:absolute;left:0;top:.62em;
width:.46em;height:.46em;background:%(accent)s;border-radius:50%%}
.cover{justify-content:center;align-items:flex-start}
.cover .stitle{font-size:clamp(28px,4.4vw,56px);max-width:92%%}
.csub{font-size:clamp(12px,1.35vw,18px);color:%(muted)s;margin-top:1.8vh}
.ckpis{display:flex;gap:3vw;margin-top:5vh;flex-wrap:wrap}
.ckpi .k{font-size:clamp(10px,1vw,13px);color:%(muted)s}
.ckpi .v{font-size:clamp(22px,3vw,40px);font-weight:700;
font-variant-numeric:tabular-nums;letter-spacing:-.03em}
.dl{display:grid;grid-template-columns:auto 1fr;gap:1.1vh 1.6vw;
font-size:clamp(12px,1.3vw,18px);align-content:center}
.dl dt{color:%(muted)s;white-space:nowrap}
.dl dd{margin:0}
table{border-collapse:collapse;width:100%%;font-size:clamp(12px,1.3vw,18px)}
th,td{padding:.9vh 1.1vw;border-bottom:1px solid %(rule)s;text-align:left}
th{color:%(muted)s;font-weight:600;white-space:nowrap}
td.num{text-align:right;font-variant-numeric:tabular-nums}
.pg{position:absolute;right:7vw;bottom:3.4vh;font-size:clamp(10px,1vw,13px);
color:%(muted)s;font-variant-numeric:tabular-nums}
.brand{position:absolute;left:7vw;bottom:3.4vh;font-size:clamp(10px,1vw,13px);
color:%(muted)s}
.hint{position:fixed;left:50%%;bottom:14px;transform:translateX(-50%%);
font-size:11px;color:#8A8F98;background:rgba(0,0,0,.55);padding:4px 12px;
border-radius:11px;transition:opacity .4s}
.hint.gone{opacity:0}
@media print{
 @page{size:landscape;margin:0}
 html,body{background:#fff}
 .slide{display:flex!important;page-break-after:always;
 width:100%%;height:100vh}
 .hint{display:none}
}
"""

DECK_JS = """
(function(){
 var s=[].slice.call(document.querySelectorAll('.slide')),i=0;
 function go(n){i=Math.max(0,Math.min(s.length-1,n));
  s.forEach(function(e,k){e.classList.toggle('on',k===i)});
  location.hash='p'+(i+1);}
 document.addEventListener('keydown',function(e){
  if(['ArrowRight','PageDown',' ','Enter'].indexOf(e.key)>=0){e.preventDefault();go(i+1)}
  if(['ArrowLeft','PageUp','Backspace'].indexOf(e.key)>=0){e.preventDefault();go(i-1)}
  if(e.key==='Home')go(0); if(e.key==='End')go(s.length-1);
 });
 document.addEventListener('click',function(e){
  if(e.target.closest('a'))return; go(e.clientX>innerWidth*0.35?i+1:i-1)});
 var m=(location.hash||'').match(/p(\\d+)/); go(m?parseInt(m[1])-1:0);
 var h=document.querySelector('.hint');
 if(h)setTimeout(function(){h.classList.add('gone')},3200);
})();
"""


def build_deck(spec: dict) -> str:
    """HTML 幻灯片。每页一个 message title —— 页标题就是这一页的结论。

    自包含，键盘/点击翻页，浏览器打印即得每页一张的 PDF。
    """
    st = resolve_style(spec)
    slides: list[str] = []
    n_body = 0

    # 封面
    kpis = spec.get("kpis") or []
    cover = [f'<div class="slide cover on">']
    cover.append(f'<h1 class="stitle">{E(spec.get("title", "数据分析"))}</h1>')
    if spec.get("subtitle"):
        cover.append(f'<div class="csub">{E(spec["subtitle"])}</div>')
    if kpis:
        cards = "".join(
            f'<div class="ckpi"><div class="k">{E(str(k.get("label","")))}</div>'
            f'<div class="v">{_fmt(k["value"]) if isinstance(k.get("value"), (int,float)) else E(str(k.get("value","")))}</div></div>'
            for k in kpis[:4])
        cover.append(f'<div class="ckpis">{cards}</div>')
    cover.append("</div>")
    slides.append("".join(cover))

    # 结论页：summary 单独成页，因为它是整份东西的主论点
    summary = spec.get("summary") or []
    if summary:
        n_body += 1
        pts = "".join(f"<li>{E(s)}</li>" for s in summary)
        slides.append(
            f'<div class="slide"><div class="skicker">核心结论</div>'
            f'<h2 class="stitle">结论与建议</h2>'
            f'<div class="sbody"><div class="scol">'
            f'<ul class="spoints">{pts}</ul></div></div>'
            f'<div class="pg">{n_body}</div></div>')

    # 内容页：一节一页；一节有多张图就拆多页，一页只讲一件事
    for sec in spec.get("sections", []):
        charts = sec.get("charts") or ([sec["chart"]] if sec.get("chart") else [])
        tables = sec.get("tables") or ([sec["table"]] if sec.get("table") else [])
        paras = [p.strip() for p in (sec.get("body") or "").split("\n\n") if p.strip()]
        blocks: list[tuple[str, Any]] = [("chart", c) for c in charts] \
            + [("table", t) for t in tables]
        if not blocks:
            blocks = [("text", None)]
        for bi, (kind, payload) in enumerate(blocks):
            n_body += 1
            head = E(sec.get("heading", ""))
            if bi > 0:
                head += "（续）"
            vis = ""
            if kind == "chart":
                vis = render_chart(payload, st).replace('<figure class="chart">', '<div>') \
                                               .replace("</figure>", "</div>")
            elif kind == "table":
                vis = render_table(payload, st)
            text_paras = paras if bi == 0 else []
            # 版式跟内容走：解读短（≤2 段）用全宽大图 + 底部脚注条——
            # 图是这一页的论证主体，解读降为脚注；文字确实多才左右分栏。
            if vis and len(text_paras) <= 2:
                foot = ""
                if text_paras:
                    foot = ('<div class="sfoot">'
                            + "".join(f"<p>{E(p)}</p>" for p in text_paras)
                            + "</div>")
                inner = (f'<div class="sbody one"><div class="scol">{vis}</div>'
                         f'</div>{foot}')
            elif vis:
                pts = ('<ul class="spoints">'
                       + "".join(f"<li>{E(p)}</li>" for p in text_paras) + "</ul>")
                inner = (f'<div class="sbody"><div class="scol">{vis}</div>'
                         f'<div class="scol">{pts}</div></div>')
            else:
                pts = ('<ul class="spoints">'
                       + "".join(f"<li>{E(p)}</li>" for p in text_paras) + "</ul>")
                inner = f'<div class="sbody"><div class="scol">{pts}</div></div>'
            slides.append(
                f'<div class="slide"><h2 class="stitle">{head}</h2>'
                f'{inner}<div class="pg">{n_body}</div></div>')

    # 口径页
    cal = spec.get("caliber") or {}
    if cal:
        n_body += 1
        items = "".join(f"<dt>{E(str(k))}</dt><dd>{E(str(v))}</dd>" for k, v in cal.items())
        slides.append(
            f'<div class="slide"><div class="skicker">附录</div>'
            f'<h2 class="stitle">口径与方法</h2>'
            f'<div class="sbody"><div class="scol"><dl class="dl">{items}</dl></div></div>'
            f'<div class="pg">{n_body}</div></div>')

    # 边界页
    bounds = spec.get("boundaries") or []
    if bounds:
        n_body += 1
        pts = "".join(f"<li>{E(b)}</li>" for b in bounds)
        slides.append(
            f'<div class="slide"><div class="skicker">附录</div>'
            f'<h2 class="stitle">边界与风险提示</h2>'
            f'<div class="sbody"><div class="scol">'
            f'<ul class="spoints">{pts}</ul></div></div>'
            f'<div class="pg">{n_body}</div></div>')

    return f"""<!doctype html>
<html lang="zh-CN"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{E(spec.get('title', '数据分析'))}</title>
<style>{DECK_CSS % st}{(st.get('extra_css') or '') % st}</style></head>
<body>{''.join(slides)}
<div class="hint">← → 翻页 ｜ 点击左右区域 ｜ 打印即得 PDF</div>
<script>{DECK_JS}</script></body></html>"""


# ── 形态三：Excel 内的报告 ──────────────────────────────────────────
def build_xlsx(spec: dict, out: Path) -> str:
    """把洞察写进 Excel：文字结论 + 原生图表 + 可下钻的数据表。

    适合读者要在 Excel 里继续动手的场景——他能改数、能重排、能加自己的列，
    图表跟着变。HTML 报告是死的，这个是活的。
    """
    try:
        from openpyxl import Workbook
        from openpyxl.chart import BarChart, LineChart, Reference
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise SystemExit("生成 xlsx 需要 openpyxl：pip install openpyxl")

    st = resolve_style(spec)
    accent = st["accent"].lstrip("#")
    muted = st["muted"].lstrip("#")
    fg = st["fg"].lstrip("#")

    wb = Workbook()
    ws = wb.active
    ws.title = "报告"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 96
    r = 2

    def put(text: str, size: int = 11, bold: bool = False, color: str = fg,
            gap: int = 1, wrap: bool = True) -> None:
        nonlocal r
        c = ws.cell(row=r, column=2, value=text)
        c.font = Font(size=size, bold=bold, color=color)
        c.alignment = Alignment(wrap_text=wrap, vertical="top")
        if wrap:
            # 粗略估行高：按 46 个中文字一行
            ws.row_dimensions[r].height = max(16, 15 * (len(text) // 46 + 1))
        r += gap

    put(spec.get("title", "数据分析报告"), size=18, bold=True, gap=1)
    if spec.get("subtitle"):
        put(spec["subtitle"], size=10, color=muted, gap=2)

    kpis = spec.get("kpis") or []
    if kpis:
        for i, k in enumerate(kpis[:6]):
            col = 2 + i
            ws.cell(row=r, column=col, value=str(k.get("label", ""))).font = \
                Font(size=9, color=muted)
            v = k.get("value")
            vc = ws.cell(row=r + 1, column=col,
                         value=v if isinstance(v, (int, float)) else str(v))
            vc.font = Font(size=16, bold=True, color=fg)
            if isinstance(v, (int, float)):
                vc.number_format = "#,##0.##"
            if k.get("note"):
                ws.cell(row=r + 2, column=col, value=str(k["note"])).font = \
                    Font(size=9, color=muted)
            if i:
                ws.column_dimensions[get_column_letter(col)].width = 18
        r += 4

    summary = spec.get("summary") or []
    if summary:
        put("核心结论", size=13, bold=True, color=accent, gap=1)
        for s in summary:
            put("· " + s, size=11)
        r += 1

    cal = spec.get("caliber") or {}
    if cal:
        put("口径", size=13, bold=True, color=accent, gap=1)
        for k, v in cal.items():
            put(f"{k}：{v}", size=10, color=muted)
        r += 1

    # 数据表放独立 sheet，图表引用它 —— 这样用户改数图会跟着变
    data_ws = wb.create_sheet("数据")
    data_ws.sheet_view.showGridLines = False
    drow = 1
    chart_n = 0

    for sec in spec.get("sections", []):
        if sec.get("heading"):
            put(sec["heading"], size=13, bold=True, gap=1)
        for para in (sec.get("body") or "").split("\n\n"):
            if para.strip():
                put(para.strip(), size=11)

        for ch in (sec.get("charts") or ([sec["chart"]] if sec.get("chart") else [])):
            labels = [str(x) for x in ch.get("labels", [])]
            if ch.get("type") == "line":
                series = ch.get("series") or [{"name": ch.get("name", ""),
                                               "values": ch.get("values", [])}]
            else:
                vals = ch.get("values", [])
                pairs = list(zip(labels, vals))
                if ch.get("sort", True):
                    pairs.sort(key=lambda t: (t[1] is None, -(t[1] or 0)))
                labels = [p[0] for p in pairs]
                series = [{"name": ch.get("caption", "值"), "values": [p[1] for p in pairs]}]
            if not labels:
                continue
            chart_n += 1
            head_row = drow
            data_ws.cell(row=drow, column=1, value=ch.get("caption", f"图{chart_n}")).font = \
                Font(bold=True)
            drow += 1
            data_ws.cell(row=drow, column=1, value="类别").font = Font(size=9, color=muted)
            for j, s in enumerate(series):
                data_ws.cell(row=drow, column=2 + j, value=str(s.get("name", f"系列{j+1}"))) \
                    .font = Font(size=9, color=muted)
            first = drow
            for i, lab in enumerate(labels):
                data_ws.cell(row=drow + 1 + i, column=1, value=lab)
                for j, s in enumerate(series):
                    v = (s.get("values") or [None] * len(labels))[i] \
                        if i < len(s.get("values") or []) else None
                    data_ws.cell(row=drow + 1 + i, column=2 + j, value=v)
            last = drow + len(labels)

            c = LineChart() if ch.get("type") == "line" else BarChart()
            if isinstance(c, BarChart):
                c.type = "bar"
                c.x_axis.scaling.min = 0     # 条形编码长度，基线必须为 0
            c.add_data(Reference(data_ws, min_col=2, max_col=1 + len(series),
                                 min_row=first, max_row=last), titles_from_data=True)
            c.set_categories(Reference(data_ws, min_col=1, min_row=first + 1, max_row=last))
            c.title = ch.get("caption", "")
            c.height, c.width = 8, 17
            if hasattr(c, "y_axis"):
                c.y_axis.majorGridlines = None
            if len(series) == 1:
                c.legend = None
            try:
                for i, sr in enumerate(c.series):
                    col = SERIES_COLORS[i % len(SERIES_COLORS)].lstrip("#")
                    if isinstance(c, LineChart):
                        sr.graphicalProperties.line.solidFill = col
                    else:
                        sr.graphicalProperties.solidFill = col
            except (AttributeError, TypeError):
                pass
            ws.add_chart(c, f"B{r}")
            r += 17
            drow = last + 2
            if ch.get("note"):
                put(ch["note"], size=9, color=muted)

        for tb in (sec.get("tables") or ([sec["table"]] if sec.get("table") else [])):
            cols = tb.get("columns") or []
            if not cols:
                continue
            if tb.get("caption"):
                put(tb["caption"], size=10, bold=True, color=muted)
            hdr = r
            thin = Side(style="thin", color="D0D0D0")
            for j, cn in enumerate(cols):
                cell = ws.cell(row=hdr, column=2 + j, value=str(cn))
                cell.font = Font(size=10, bold=True, color=muted)
                cell.border = Border(bottom=thin)
            for i, row in enumerate(tb.get("rows") or []):
                for j, v in enumerate(row):
                    cell = ws.cell(row=hdr + 1 + i, column=2 + j, value=v)
                    cell.font = Font(size=10)
                    if isinstance(v, (int, float)) and not isinstance(v, bool):
                        cell.number_format = "#,##0.##"
                        cell.alignment = Alignment(horizontal="right")
            for j in range(1, len(cols)):
                ws.column_dimensions[get_column_letter(2 + j)].width = 16
            r = hdr + len(tb.get("rows") or []) + 2

    bounds = spec.get("boundaries") or []
    if bounds:
        r += 1
        put("这份分析的边界", size=12, bold=True, color=accent, gap=1)
        for b in bounds:
            put("· " + b, size=10, color=muted)

    ws.freeze_panes = "A2"
    wb.save(out)
    return f"Excel 内报告，图表引用「数据」sheet —— 改数图会跟着变"


# ── 形态四：六页纸文档 ──────────────────────────────────────────────
def _dx_esc(s: str) -> str:
    return (str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))


# Word 的字体只能写死一个名字，没有 CSS 那种 fallback 链——
# 所以这里必须选**Microsoft Office 在 Windows 和 macOS 上都随装**的字体。
# 写 PingFang SC / Helvetica 这类 macOS 独有字体，在 Windows 上会静默回退成
# 主题字体（通常是宋体），而且生成端一个警告都不会报。
# 备选同样安全：DengXian(等线) / SimSun(宋体) / SimHei(黑体)。
# 改这里之前先跑 verify_docx.py，它会拦住平台绑定的字体。
_DX_CJK = "Microsoft YaHei"
_DX_LATIN = "Microsoft YaHei"


_EMU_PER_PX = 9525            # 96 dpi
_DOCX_TEXT_W_EMU = 9354 * 635  # 版心宽：A4 减页边距（twip → EMU）


def _png_dims(data: bytes) -> tuple[int, int]:
    """PNG 头里的宽高。IHDR 固定在第 16–24 字节，纯标准库。"""
    import struct
    return struct.unpack(">II", data[16:24])


def _rasterize_svgs(svgs: list[str], st: dict, css_w: int = 1100) -> list[bytes] | None:
    """内联 SVG → PNG（2x）。依赖 playwright——它已经是第 7 步渲染闸门的
    依赖，不新增环境要求；起不来就返回 None，调用方降级为数据表并明说。"""
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        return None
    blocks = "".join(
        f'<div id="c{i}" style="width:{css_w}px;background:{st["bg"]};'
        f'padding:14px 10px">{s}</div>' for i, s in enumerate(svgs))
    doc = (f'<!doctype html><meta charset="utf-8"><body style="margin:0;'
           f'font-family:{st["b_font"]}">{blocks}</body>')
    try:
        with sync_playwright() as p:
            b = p.chromium.launch()
            pg = b.new_page(viewport={"width": css_w + 40, "height": 800},
                            device_scale_factor=2)
            pg.set_content(doc)
            pg.wait_for_timeout(120)
            shots = [pg.query_selector(f"#c{i}").screenshot(type="png")
                     for i in range(len(svgs))]
            b.close()
        return shots
    except Exception:
        return None


def _docx_image(rid: str, n: int, px_w: int, px_h: int) -> str:
    """内嵌图片的 OOXML。宽度取版心与实际宽度（2x 渲染折半）的较小者，
    永远不超版心——verify_docx.py 的第 ④ 项查的就是这个。"""
    cx = min(px_w * _EMU_PER_PX // 2, _DOCX_TEXT_W_EMU)
    cy = int(cx * px_h / max(px_w, 1))
    return (
        f'<w:p><w:pPr><w:spacing w:after="80"/></w:pPr><w:r><w:drawing>'
        f'<wp:inline distT="0" distB="0" distL="0" distR="0">'
        f'<wp:extent cx="{cx}" cy="{cy}"/>'
        f'<wp:docPr id="{n}" name="chart{n}"/>'
        f'<a:graphic><a:graphicData '
        f'uri="http://schemas.openxmlformats.org/drawingml/2006/picture">'
        f'<pic:pic>'
        f'<pic:nvPicPr><pic:cNvPr id="{n}" name="chart{n}.png"/>'
        f'<pic:cNvPicPr/></pic:nvPicPr>'
        f'<pic:blipFill><a:blip r:embed="{rid}"/>'
        f'<a:stretch><a:fillRect/></a:stretch></pic:blipFill>'
        f'<pic:spPr><a:xfrm><a:off x="0" y="0"/><a:ext cx="{cx}" cy="{cy}"/></a:xfrm>'
        f'<a:prstGeom prst="rect"><a:avLst/></a:prstGeom></pic:spPr>'
        f'</pic:pic></a:graphicData></a:graphic></wp:inline></w:drawing></w:r></w:p>')


def build_docx(spec: dict, out: Path) -> str:
    """Word 文档，按 Amazon six-pager 的路子：叙述体，不是幻灯片。

    six-pager 的精髓不是「六页」，是**用完整段落写**——
    bullet 允许你把没想清楚的东西并列摆着蒙混过关，
    完整句子会逼你写出因果和取舍。写不清楚，就是没想清楚。

    所以这里把 summary 也渲染成段落而非项目符号。

    图表以 PNG 内嵌正文（经 playwright 渲染，和 HTML 版长得一样），
    数据表进附录供引用——图和表不互替：给 VP 的 Word 零可视化站不住，
    这是四场压力测试一致的判决。环境缺 playwright 时降级为纯数据表，
    并在文档里明说，不静默。

    手写 OOXML，纯标准库，不需要 python-docx。
    """
    import zipfile

    st = resolve_style(spec)
    body: list[str] = []

    def para(text: str, style: str = "Body", ) -> None:
        body.append(
            f'<w:p><w:pPr><w:pStyle w:val="{style}"/></w:pPr>'
            f'<w:r><w:t xml:space="preserve">{_dx_esc(text)}</w:t></w:r></w:p>')

    def table(cols: list, rows: list) -> None:
        w = int(9360 / max(len(cols), 1))
        tr = ['<w:tbl><w:tblPr><w:tblStyle w:val="TableGrid"/>'
              '<w:tblW w:w="9360" w:type="dxa"/></w:tblPr>']
        tr.append("<w:tblGrid>" + "".join(f'<w:gridCol w:w="{w}"/>' for _ in cols)
                  + "</w:tblGrid>")
        # 表头行：tblHeader 让它跨页重复，cantSplit 防止被拦腰截断。
        # 缺这两个属性时本机看着完好，跨页后读者看不到列名——
        # verify_docx.py 的第 ⑤ 项查的就是它，生成端必须自己先过自己的闸门。
        tr.append('<w:tr><w:trPr><w:tblHeader/><w:cantSplit/></w:trPr>' + "".join(
            f'<w:tc><w:tcPr><w:tcW w:w="{w}" w:type="dxa"/></w:tcPr>'
            f'<w:p><w:pPr><w:pStyle w:val="Body"/></w:pPr><w:r><w:rPr><w:b/></w:rPr>'
            f'<w:t>{_dx_esc(c)}</w:t></w:r></w:p></w:tc>' for c in cols) + "</w:tr>")
        for row in rows:
            cells = []
            for v in row:
                txt = _fmt(v) if isinstance(v, (int, float)) and not isinstance(v, bool) \
                    else str(v)
                cells.append(f'<w:tc><w:tcPr><w:tcW w:w="{w}" w:type="dxa"/></w:tcPr>'
                             f'<w:p><w:pPr><w:pStyle w:val="Body"/></w:pPr>'
                             f'<w:r><w:t>{_dx_esc(txt)}</w:t></w:r></w:p></w:tc>')
            tr.append("<w:tr>" + "".join(cells) + "</w:tr>")
        tr.append("</w:tbl>")
        body.append("".join(tr))
        para("")

    # 图表统一先渲一遍：一次浏览器会话渲完所有图，失败就整体降级
    charts_all = [ch for sec in spec.get("sections", [])
                  for ch in (sec.get("charts") or
                             ([sec["chart"]] if sec.get("chart") else []))]
    chart_svgs = [(ch, _chart_svg(ch, st)) for ch in charts_all]
    chart_svgs = [(ch, s) for ch, s in chart_svgs if s]
    rendered = _rasterize_svgs([s for _, s in chart_svgs], st) if chart_svgs else []
    png_map: dict[int, bytes] = {}
    if rendered:
        for (ch, _), png in zip(chart_svgs, rendered):
            png_map[id(ch)] = png
    degraded = bool(chart_svgs) and not rendered
    media: list[tuple[str, bytes]] = []
    img_no = 0

    para(spec.get("title", "数据分析报告"), "Title")
    if spec.get("subtitle"):
        para(spec["subtitle"], "Sub")

    summary = spec.get("summary") or []
    if summary:
        para("摘要", "H1")
        # 刻意合成段落而不是逐条列出：six-pager 不用 bullet
        para("".join(s.rstrip("。.") + "。" for s in summary))

    cal = spec.get("caliber") or {}
    if cal:
        para("口径", "H1")
        para("；".join(f"{k}：{v}" for k, v in cal.items()) + "。")

    if degraded:
        # 降级要说出来，不能静默——读者该知道这份 Word 缺了什么、去哪补
        para("说明：本环境缺少 playwright，图表未能渲染嵌入，仅以数据表"
             "形式列于附录；完整图表见同名 HTML 报告。", "Note")

    appendix: list[tuple[str, list, list]] = []
    for sec in spec.get("sections", []):
        if sec.get("heading"):
            para(sec["heading"], "H1")
        for p in (sec.get("body") or "").split("\n\n"):
            if p.strip():
                para(p.strip())
        for ch in (sec.get("charts") or ([sec["chart"]] if sec.get("chart") else [])):
            png = png_map.get(id(ch))
            if png is not None:
                img_no += 1
                if ch.get("caption"):
                    para(f"图 {img_no}｜{ch['caption']}", "Caption")
                media.append((f"chart{img_no}.png", png))
                body.append(_docx_image(f"rIdImg{img_no}", img_no, *_png_dims(png)))
            labels = [str(x) for x in ch.get("labels", [])]
            if not labels:
                continue
            if ch.get("type") == "line":
                series = ch.get("series") or [{"name": ch.get("name", "值"),
                                               "values": ch.get("values", [])}]
                cols = ["项"] + [str(s.get("name", "值")) for s in series]
                rows = [[labels[i]] + [(s.get("values") or [None])[i]
                                       if i < len(s.get("values") or []) else None
                                       for s in series] for i in range(len(labels))]
            else:
                vals = ch.get("values", [])
                pairs = sorted(zip(labels, vals), key=lambda t: -(t[1] or 0)) \
                    if ch.get("sort", True) else list(zip(labels, vals))
                cols = ["项", str(ch.get("caption", "值"))]
                rows = [[a, b] for a, b in pairs]
            appendix.append((ch.get("caption", "数据"), cols, rows))
            if ch.get("note"):
                para(ch["note"], "Note")
        for tb in (sec.get("tables") or ([sec["table"]] if sec.get("table") else [])):
            if tb.get("columns"):
                if tb.get("caption"):
                    para(tb["caption"], "H2")
                table(tb["columns"], tb.get("rows") or [])

    bounds = spec.get("boundaries") or []
    if bounds:
        para("这份分析的边界", "H1")
        para("".join(b.rstrip("。.") + "。" for b in bounds))

    if appendix:
        para("附录：数据", "H1")
        for cap, cols, rows in appendix:
            para(cap, "H2")
            table(cols, rows)

    doc = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
           '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main" '
           'xmlns:wp="http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing" '
           'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" '
           'xmlns:pic="http://schemas.openxmlformats.org/drawingml/2006/picture" '
           'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
           f'<w:body>{"".join(body)}'
           '<w:sectPr><w:pgSz w:w="11906" w:h="16838"/>'
           '<w:pgMar w:top="1418" w:right="1276" w:bottom="1418" w:left="1276"/>'
           '</w:sectPr></w:body></w:document>')

    def _style(sid: str, name: str, size: int, bold: bool, before: int,
               after: int, color: str = "000000") -> str:
        return (f'<w:style w:type="paragraph" w:styleId="{sid}">'
                f'<w:name w:val="{name}"/><w:qFormat/><w:pPr>'
                f'<w:spacing w:before="{before}" w:after="{after}" w:line="312" '
                f'w:lineRule="auto"/></w:pPr><w:rPr>'
                f'<w:rFonts w:ascii="{_DX_LATIN}" w:hAnsi="{_DX_LATIN}" '
                f'w:eastAsia="{_DX_CJK}" w:cs="{_DX_LATIN}"/>'
                f'{"<w:b/>" if bold else ""}<w:color w:val="{color}"/>'
                f'<w:sz w:val="{size*2}"/><w:szCs w:val="{size*2}"/></w:rPr></w:style>')

    styles = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
              '<w:styles xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
              + _style("Title", "Title", 20, True, 0, 160)
              + _style("Sub", "Subtitle", 10, False, 0, 320, "6B6B6B")
              + _style("H1", "heading 1", 14, True, 360, 140)
              + _style("H2", "heading 2", 11, True, 240, 100)
              + _style("Body", "Body Text", 11, False, 0, 140)
              + _style("Caption", "caption", 10, True, 200, 60)
              + _style("Note", "Note", 9, False, 20, 160, "6B6B6B")
              + '<w:style w:type="table" w:styleId="TableGrid"><w:name w:val="Table Grid"/>'
                '<w:tblPr><w:tblBorders>'
                + "".join(f'<w:{e} w:val="single" w:sz="4" w:color="CCCCCC"/>'
                          for e in ("top", "left", "bottom", "right",
                                    "insideH", "insideV"))
                + '</w:tblBorders></w:tblPr></w:style>'
              + "</w:styles>")

    ct = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
          '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
          '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
          '<Default Extension="xml" ContentType="application/xml"/>'
          '<Default Extension="png" ContentType="image/png"/>'
          '<Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>'
          '<Override PartName="/word/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.styles+xml"/>'
          "</Types>")
    rels = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>'
            "</Relationships>")
    img_rels = "".join(
        f'<Relationship Id="rIdImg{i}" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image" '
        f'Target="media/{name}"/>'
        for i, (name, _) in enumerate(media, 1))
    drels = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
             '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
             '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>'
             f"{img_rels}</Relationships>")

    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("[Content_Types].xml", ct)
        z.writestr("_rels/.rels", rels)
        z.writestr("word/document.xml", doc)
        z.writestr("word/_rels/document.xml.rels", drels)
        z.writestr("word/styles.xml", styles)
        for name, data in media:
            z.writestr(f"word/media/{name}", data)

    if media:
        return f"六页纸叙述体文档，{len(media)} 张图表内嵌，数据表进附录"
    return "六页纸叙述体文档，图表数据进附录" + \
        ("（缺 playwright，图未嵌入）" if degraded else "")


def audit_spec(spec: dict) -> list[str]:
    """产出前自检：哪些地方本该是图，却写成了字或摆成了表。

    这些都是机器能判的客观量——字数、图表数、表格的行列形状。
    「这个洞察深不深」机器判不了，但「这一屏全是字」它判得了。
    """
    tips: list[str] = []
    text_len = 0
    charts = 0
    tables: list[dict] = []
    for sec in spec.get("sections", []):
        text_len += len(sec.get("heading", "")) + len(sec.get("body", ""))
        for sub in sec.get("subsections", []):
            text_len += len(sub.get("body", ""))
        charts += len(sec.get("charts") or ([1] if sec.get("chart") else []))
        tables += (sec.get("tables") or ([sec["table"]] if sec.get("table") else []))
    text_len += sum(len(s) for s in (spec.get("summary") or []))

    # 一张图约顶 800 字。差得太远说明该看的东西都埋在段落里了。
    if text_len > 1500 and charts < text_len / 800:
        want = int(text_len / 800)
        tips.append(
            f"正文约 {text_len} 字，只有 {charts} 张图。这个密度读者会跳着看。"
            f"按一张图顶 800 字算，至少该有 {want} 张。"
            f"回去找：哪几段其实在描述一个趋势、一个排名、或一个构成？那些都该是图。")

    for i, tb in enumerate(tables, 1):
        rows = tb.get("rows") or []
        cols = tb.get("columns") or []
        if len(rows) < 5 or not cols:
            continue
        numeric_cols = 0
        for j in range(len(cols)):
            vals = [r[j] for r in rows if j < len(r)]
            if vals and all(isinstance(v, (int, float)) and not isinstance(v, bool)
                            for v in vals):
                numeric_cols += 1
        cap = tb.get("caption") or f"第 {i} 个表"
        if numeric_cols == 1:
            tips.append(
                f"「{cap}」{len(rows)} 行 × 1 个数值列 —— 这是排名，画成条形图读者"
                f"一眼能看出差距。表格适合「要查具体数值」，不适合「要看谁高谁低」。")
        elif numeric_cols >= 1 and len(rows) >= 8:
            has_neg = any(isinstance(v, (int, float)) and v < 0
                          for r in rows for v in r
                          if isinstance(v, (int, float)) and not isinstance(v, bool))
            if has_neg:
                tips.append(
                    f"「{cap}」{len(rows)} 行且含正负值 —— 如果它在解释「总量为什么"
                    f"变了」，瀑布图（`type: waterfall`）比表格清楚一个数量级。")
    return tips


TEMPLATE = {
    "title": "标题写结论，不写主题（「华东贡献六成销售额但增速垫底」而不是「销售额分析」）",
    "subtitle": "数据范围 / 制表日期",
    "style": "report",
    "_style_note": "report=默认业务报告 | finance=财务审计类 | minimal=内部快看。"
                   "你按内容自己选，不要拿去问用户。",
    "kpis": [
        {"label": "上半年销售额", "value": 14688217, "note": "较去年同期 +12.4%", "delta": 1}
    ],
    "summary": [
        "第一条结论（能改变某个决定的那种，不是描述）",
        "第二条结论"
    ],
    "caliber": {
        "时间窗口": "2026-01-01 至 2026-06-30，按支付时间",
        "口径": "含税，不含运费；已排除汇总行；未去重",
        "数据来源": "销售明细.xlsx"
    },
    "sections": [
        {
            "heading": "小标题也写结论",
            "body": "段落。空行分段。\n\n第二段。",
            "chart": {
                "type": "bar",
                "caption": "图表标题写结论",
                "labels": ["华东", "华南", "华北"],
                "values": [6497500, 4556300, 2654317],
                "sort": True,
                "highlight": "华东",
                "note": "可选的图下注解：只写图的意义，不复述图里已有的数字"
            }
        },
        {
            "heading": "趋势",
            "body": "",
            "chart": {
                "type": "line",
                "caption": "三月起连续下滑",
                "labels": ["1月", "2月", "3月", "4月", "5月", "6月"],
                "series": [{"name": "销售额", "values": [120, 135, 128, 119, 108, 96]}],
                "incomplete_last": False,
                "note": "incomplete_last=true 会把最后一段画成虚线——"
                        "本期没走完却画成实线，是最常见的误导"
            },
            "table": {
                "caption": "可选表格",
                "columns": ["门店", "销售额", "客户数"],
                "rows": [["朝阳店", 1234567, 1203]]
            }
        }
    ],
    "boundaries": [
        "数据只覆盖 X 到 Y，结论在此之外不成立",
        "哪个结论最脆弱、换什么口径会翻盘",
        "哪些数字没有独立校验来源"
    ],
    "footer": "生成于 … ｜ 口径见上"
}


def main() -> None:
    ap = argparse.ArgumentParser(description="把洞察渲染成自包含 HTML 报告")
    ap.add_argument("--spec", help="洞察 spec 的 JSON 文件路径")
    ap.add_argument("--out", default=None, help="输出路径，默认按 --format 定后缀")
    ap.add_argument("--format", default="html",
                    choices=["html", "deck", "xlsx", "docx"],
                    help="html=网页报告 deck=幻灯片 xlsx=Excel内报告 docx=六页纸文档")
    ap.add_argument("--template", action="store_true", help="打印一份 spec 模板")
    ap.add_argument("--styles", action="store_true", help="列出可用风格")
    a = ap.parse_args()

    if a.styles:
        for k, v in STYLES.items():
            print(f"{k:10} {v['desc']}")
        print("\n按内容自己选，不要拿去问用户——报告不是设计作品。")
        return
    if a.template:
        print(json.dumps(TEMPLATE, ensure_ascii=False, indent=2))
        return
    if not a.spec:
        sys.exit("需要 --spec <json文件>，或用 --template 拿一份模板")

    sp = Path(a.spec).expanduser()
    if not sp.exists():
        sys.exit(f"文件不存在：{sp}")
    try:
        spec = json.loads(sp.read_text(encoding="utf-8"))
    except json.JSONDecodeError as e:
        sys.exit(f"spec 不是合法 JSON：{e}")

    ext = {"html": ".html", "deck": ".deck.html", "xlsx": ".xlsx", "docx": ".docx"}[a.format]
    out = Path(a.out).expanduser() if a.out else Path("report" + ext)

    if a.format == "html":
        out.write_text(build_html(spec), encoding="utf-8")
        note = "自包含，无外部资源"
    elif a.format == "deck":
        out.write_text(build_deck(spec), encoding="utf-8")
        note = "自包含幻灯片，← → 翻页，浏览器打印即得每页一张的 PDF"
    elif a.format == "xlsx":
        note = build_xlsx(spec, out)
    else:
        note = build_docx(spec, out)

    size = out.stat().st_size / 1024
    print(f"已生成：{out}  ({size:.1f} KB，{note})")

    # deck 和 html 都靠浏览器渲染，后缀不对浏览器会当纯文本显示，
    # 而且不会有任何报错——只是打开后一片源码。
    if a.format in ("html", "deck") and out.suffix.lower() not in (".html", ".htm"):
        print(f"  ⚠ 文件名没有 .html 后缀，浏览器会把它当纯文本显示。"
              f"建议改成 {out.with_suffix('.html').name}")
    missing = []
    if not spec.get("caliber"):
        missing.append("caliber（口径声明）—— 没有口径的数字无法被检验")
    if not spec.get("boundaries"):
        missing.append("boundaries（分析边界）—— 说出自身弱点的报告才可信")
    if not spec.get("summary"):
        missing.append("summary（结论先行）—— 读者不该自己去找结论")
    for m in missing:
        print(f"  ⚠ 缺 {m}")
    for t in audit_spec(spec):
        print(f"  → {t}")


if __name__ == "__main__":
    main()
